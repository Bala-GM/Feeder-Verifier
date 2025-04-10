import pandas as pd
import numpy as np
import os
from os import getcwd
import csv as csv
from io import StringIO
import re
import datetime
from io import BytesIO
import time
import sys
from datetime import datetime
import shutil
import linecache
import ast
import tkinter as tk
from tkinter import *
from tkinter import scrolledtext, messagebox
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox 
from tkinter.filedialog import askopenfile
import subprocess
import threading
import urllib
import urllib.parse
from sqlite3 import dbapi2 as sqlite
import sqlite3
import lxml
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlsxwriter
import seaborn as ssn
import matplotlib
import matplotlib as mlp
from matplotlib import pyplot as plt
import xlwings as xw
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pyxcelframe
from openpyxl import load_workbook
from pyxcelframe import copy_cell_style
from pyxcelframe import insert_frame
from pyxcelframe import insert_columns
from pyxcelframe import sheet_to_sheet
from pyxcelframe import column_last_row
import sqlalchemy #pip install SQLAlchemy
from sqlalchemy import create_engine
import sqlalchemy_access as sq_a #pip install sqlalchemy-access
import sqlalchemy_pyodbc_mssql as sqlalchemy #pip install sqlalchemy-pyodbc-mssql
from flask_sqlalchemy import SQLAlchemy #pip install Flask-SQLAlchemy
import pyodbc
import pypyodbc #pip install pypyodbc
import odbc
import psycopg2 #pip install psycopg2
import mysql.connector as sql #pip install mysql-connector-python
import MySQLdb as sql #pip install mysqlclient
from plyer import notification #pip install plyer
import xlrd
import xml.etree.ElementTree as ET
from openpyxl.utils.exceptions import InvalidFileException
import psutil


# Program B: FeederSetup: V-2.6.3 X  Manipulator PY V-9.5.6 APR|07|04|2025 deprecated
#def program_B():
print("\033[92;40mFeeder-Verfier PY TK V-1.0.1 APR|08|04|2025 \033[0m \033[1;34;40mSYRMA\033[0m \033[1;36;40mSGS\033[0m")
print('\n')
print("\033[92;40mB. FeederSetup: V-2.6.3-X \033[1;31;40mManipulator PY V-9.5.6 APR|07|04|2025 deprecated\033[0m") #89P13
time.sleep(3)
print('\n')
print("BSD 2-Clause License Copyright (c) 2025, Bala Ganesh")
print('\n')
print("\033[32;4m*******Feeder and BOM data Verification Version--PY_V-2.6.3-X interface_TK/APR|07|04|2025-89P13*******\033[0m")
print('\n')
time.sleep(3)

# Get the current date and time
current_datetime = datetime.now()

# Format the current date and time as a string
#formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
# Format the date and time in a 12-hour clock with AM/PM
formatted_datetime = current_datetime.strftime("%Y-%m-%d %I:%M:%S %p")

print('\n')

# Print the formatted date and time
print(f"\033[31mCurrent Date and Time: {formatted_datetime}\033[0m")

print('\n')

'''print(f"Current Year: {current_datetime.year}")
print(f"Current Month: {current_datetime.month}")
print(f"Current Day: {current_datetime.day}")
print(f"Current Hour: {current_datetime.hour}")
print(f"Current Minute: {current_datetime.minute}")
print(f"Current Second: {current_datetime.second}")'''

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******CYCLE TIME*******\033[0m")
######################################################################################################

#CycleTime

# Function to find CycleTime.xml files
def find_cycle_time_files(root_directory):
    cycle_time_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "CycleTime.xml":
                cycle_time_files.append(os.path.join(root, file))
    return cycle_time_files

# Function to find the setup description from file path
def find_setup_description(file_path):
    parts = file_path.split(";")
    setup_description = parts[-2].strip()
    return setup_description

# Function to rename CycleTime.xml files with setup descriptions
def rename_cycle_time_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "CycleTime_TL1.xml"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "CycleTime_BL1.xml"
    elif "[Top] Line2" in setup_description:
        new_file_name = "CycleTime_TL2.xml"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "CycleTime_BL2.xml"
    elif "[Top] Line3" in setup_description:
        new_file_name = "CycleTime_TL3.xml"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "CycleTime_BL3.xml"
    elif "[Top] Line5" in setup_description:
        new_file_name = "CycleTime_TL5.xml"
    elif "[Bottom] Line5" in setup_description:
        new_file_name = "CycleTime_BL5.xml"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "CycleTime_TL4C.xml"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "CycleTime_BL4C.xml"
    else:
        return
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print('\n')
    print(f"CycleTime.xml renamed to: {new_path}")
    print('\n')
    return new_path

# Function to convert XML to CSV
def xml_to_csv(xml_file, csv_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Open CSV file in write mode with newline='' to prevent extra newlines
    with open(csv_file, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)

        # Write header row
        header = []
        for child in root[0]:
            header.append(child.tag)
        writer.writerow(header)

        # Write data rows
        for elem in root:
            row = []
            for child in elem:
                row.append(child.text)
            writer.writerow(row)

    print(f"Converted {xml_file} to CSV: {csv_file}")
    print('\n')
    return csv_file

# Define the root directory
root_directory = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\FeederSetup"

# Find CycleTime.xml files
cycle_time_files = find_cycle_time_files(root_directory)

# Rename CycleTime.xml files with setup descriptions, convert to CSV, and move them
for file_path in cycle_time_files:
    new_path = rename_cycle_time_with_description(file_path)
    if new_path:
        csv_file_path = xml_to_csv(new_path, os.path.splitext(new_path)[0] + ".csv")
        # Create the CycleTime directory if it doesn't exist
        cycle_time_dir = os.path.join(os.path.dirname(root_directory), "CycleTime")
        if not os.path.exists(cycle_time_dir):
            os.makedirs(cycle_time_dir)
        # Move the CSV file to the CycleTime directory
        shutil.copy(csv_file_path, os.path.join(cycle_time_dir, os.path.basename(csv_file_path)))

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******Feeder Setup*******\033[0m")
######################################################################################################

# FeederSetup

def find_feeder_setup_files(root_directory):
    feeder_setup_files = []
    for root, dirs, files in os.walk(root_directory):
        for file in files:
            if file == "FeederSetup.csv":
                feeder_setup_files.append(os.path.join(root, file))
    return feeder_setup_files

def rename_feeder_setup_with_description(file_path):
    setup_description = find_setup_description(file_path)
    if "[Top] Line1" in setup_description:
        new_file_name = "FeederSetup_TL1.csv"
    elif "[Bottom] Line1" in setup_description:
        new_file_name = "FeederSetup_BL1.csv"
    elif "[Top] Line2" in setup_description:
        new_file_name = "FeederSetup_TL2.csv"
    elif "[Bottom] Line2" in setup_description:
        new_file_name = "FeederSetup_BL2.csv"
    elif "[Top] Line3" in setup_description:
        new_file_name = "FeederSetup_TL3.csv"
    elif "[Bottom] Line3" in setup_description:
        new_file_name = "FeederSetup_BL3.csv"
    elif "[Top] Line5" in setup_description:
        new_file_name = "FeederSetup_TL5.csv"
    elif "[Bottom] Line5" in setup_description:
        new_file_name = "FeederSetup_BL5.csv"
    elif "[Top] Line-1" in setup_description:
        new_file_name = "FeederSetup_TL4C.csv"
    elif "[Bottom] Line-1" in setup_description:
        new_file_name = "FeederSetup_BL4C.csv"
    else:
        return
    
    new_path = os.path.join(os.path.dirname(file_path), new_file_name)
    os.rename(file_path, new_path)
    print('\n')
    print(f"FeederSetup.csv renamed to: {new_path}")
    print('\n')
    return new_path

# Find FeederSetup.csv files
feeder_setup_files = find_feeder_setup_files(root_directory)

# Rename FeederSetup.csv files with setup descriptions and move them
for file_path in feeder_setup_files:
    new_path = rename_feeder_setup_with_description(file_path)
    if new_path:
        # Create the FeederSetup directory if it doesn't exist
        feeder_setup_dir = os.path.join(os.path.dirname(root_directory), "FeederSetup")
        if not os.path.exists(feeder_setup_dir):
            os.makedirs(feeder_setup_dir)
        # Move the renamed file to the FeederSetup directory
        shutil.copy(new_path, os.path.join(feeder_setup_dir, os.path.basename(new_path)))

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******BOTTOM AND TOP PRPGRAM FOLDER NAME FETCH*******\033[0m")
######################################################################################################
    #BOTTOM ADN TOP FOLDER NAME FETCH
######################################################################################################

def breakdown_folder_name(folder_name):
    parts = folder_name.split(';')
    breakdown = {
        'Identifier': parts[0],
        'Version': parts[1],
        'Position': parts[2],
        'Secondary Identifier': parts[3],
        'Production Stage': parts[4],
        'Setup Details': parts[5]
    }
    return breakdown

def fetch_folder_data_and_save(root_directory, output_file):
    # Get a list of all items in the root directory
    all_items = os.listdir(root_directory)
    
    # Filter out only directories
    folder_names = [item for item in all_items if os.path.isdir(os.path.join(root_directory, item))]
    
    # Initialize an empty list to store folder data
    folder_data = []
    
    # Iterate over each folder name
    for folder_name in folder_names:
        # Break down the folder name
        breakdown = breakdown_folder_name(folder_name)
        # Append folder data to the list
        folder_data.append(breakdown)
    
    # Create a DataFrame from the folder data
    df = pd.DataFrame(folder_data)
    
    # Save the final DataFrame to a CSV file
    df.to_csv(output_file, index=False)

    # Print a message indicating the file has been saved
    print('\n')
    print(f"Folder data saved to {output_file}")
    print('\n')

    # Filter the DataFrame based on Production Stage for Bottom Line
    bottom_line_df = df[df['Production Stage'].str.startswith('[Bottom]')]
    
    # Filter the DataFrame based on Production Stage for Top Line
    top_line_df = df[df['Production Stage'].str.startswith('[Top]')]

    # Change directory to save Bottom line data
    os.chdir("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime")
    bottom_line_df.to_csv('Bottom_Line_Data.csv', index=False)
    print("Bottom line data saved to Bottom_Line_Data.csv")
    print('\n')
        
    # Change directory to save Top line data
    os.chdir("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime")
    top_line_df.to_csv('Top_Line_Data.csv', index=False)
    print("Top line data saved to Top_Line_Data.csv")
    print('\n')

# Example root directory and output file
Chd=os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')
root_directory = "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup"
output_file = 'Breakfile.csv'

#Fetch folder data and save to CSV file
fetch_folder_data_and_save(root_directory, output_file)

#Delet the BreakFile    
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')
if os.path.exists("Breakfile.csv"):
        os.remove("Breakfile.csv")
else:
    print("The file does not exist")

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******LINE ROW CHECK FOR BOT AND TOP CYCLETIME*******\033[0m")
######################################################################################################
    #CycleTime_BL1 ROW ADJUSTMENT
######################################################################################################
#conversion
        
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first NXT&AIMEX[4:] rows for old Line1 for new AIMEXIII Line1 [2:]

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_BL1.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_BL1.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_TL1 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first NXT&AIMEX[4:] rows for old Line1 for new AIMEXIII Line1 [2:]

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_TL1.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_TL1.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_BL2 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[4:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_BL2.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_BL2.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_TL2 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[4:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_TL2.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_TL2.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_BL3 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_BL3.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_BL3.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_TL3 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_TL3.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_TL3.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_BL5 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_BL5.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_BL5.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_TL5 ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[2:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_TL5.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_TL5.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_BL4C ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[3:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_BL4C.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_BL4C.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)

######################################################################################################
    #CycleTime_TL4C ROW ADJUSTMENT
######################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

def add_header_to_csv(input_file, output_file, new_header):
    # Check if the input file exists
    if not os.path.exists(input_file):
        print('\n')
        print("Input file not found:", input_file)
        return

    # Read the existing contents of the CSV file, skipping the first 4 rows
    with open(input_file, 'r', newline='') as f:
        reader = csv.reader(f)
        data = list(reader)[3:]  # Skip the first 4 rows

    # Insert the new header at the beginning of the data
    data.insert(0, new_header)

    # Write the data with the new header to the output file
    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)

# Get the current working directory
cwd = os.getcwd()
print('\n')
print("Current working directory:", cwd)

# Define the input and output file paths
input_file = "CycleTime_TL4C.csv"  # Make sure the file exists in the current directory
output_file = "CycleTime_TL4C.csv"

# New header to be added
new_header = ["LineName", "ModuleName", "ModuleNo", "Lane", "GntaryNo", "CycleTime", "Placing Time", "Loading Time", "Qty", "Average"]

# Call the function to add the new header
add_header_to_csv(input_file, output_file, new_header)
print('\n')
print("New header added to the CSV file:", output_file)
print('\n')

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******RESHAPING CYCLETIME DATA*******\033[0m")
######################################################################################################
######################################################################################################
    #Separate the Perticular Data start here
######################################################################################################
######################################################################################################

# Change directory Line 1 CycleTime Top
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_TL1.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'AIMEXIII-1':'AIMEXIII_1'}, inplace = True)
    pivot_df.rename(columns = {'AIMEXIII-2':'AIMEXIII_2'}, inplace = True)
    pivot_df.rename(columns = {'AIMEXIII-3':'AIMEXIII_3'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'AIMEXIII_1', 'AIMEXIII_2', 'AIMEXIII_3', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_TL1.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_TL1.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 1 CycleTime Bottom
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_BL1.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'AIMEXIII-1':'AIMEXIII_1'}, inplace = True)
    pivot_df.rename(columns = {'AIMEXIII-2':'AIMEXIII_2'}, inplace = True)
    pivot_df.rename(columns = {'AIMEXIII-3':'AIMEXIII_3'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'AIMEXIII_1', 'AIMEXIII_2', 'AIMEXIII_3', 'Lane', 'Qty']]

    # Display the pivot DataFrame
    print('\n')
    print("Pivot Bottom", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_BL1.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_BL1.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 2 CycleTime Top
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_TL2.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'AIMEX-IIIC_1-1':'AIMEX-IIIC_1'}, inplace = True)
    pivot_df.rename(columns = {'AIMEX-IIIC_2-1':'AIMEX-IIIC_2'}, inplace = True)
    pivot_df.rename(columns = {'AIMEX-IIIC_3-1':'AIMEX-IIIC_3'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'AIMEX-IIIC_1', 'AIMEX-IIIC_2', 'AIMEX-IIIC_3', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_TL2.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_TL2.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 2 CycleTime Bottom
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_BL2.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'AIMEX-IIIC_1-1':'AIMEX-IIIC_1'}, inplace = True)
    pivot_df.rename(columns = {'AIMEX-IIIC_2-1':'AIMEX-IIIC_2'}, inplace = True)
    pivot_df.rename(columns = {'AIMEX-IIIC_3-1':'AIMEX-IIIC_3'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'AIMEX-IIIC_1', 'AIMEX-IIIC_2', 'AIMEX-IIIC_3', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_BL2.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_BL2.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 3 CycleTime Top
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_TL3.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXTI_MC1-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-6':'NXT6'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_TL3.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_TL3.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 3 CycleTime Bottom
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_BL3.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXTI_MC1-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXTI_MC1-6':'NXT6'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_BL3.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_BL3.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 5 CycleTime Top
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_TL5.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXTIII-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-6':'NXT6'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_TL5.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_TL5.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line 5 CycleTime Bottom
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_BL5.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXTIII-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXTIII-6':'NXT6'}, inplace = True)


    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_BL5.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_BL5.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line-1 4C  CycleTime Top
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_TL4C.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXT1-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-6':'NXT6'}, inplace = True)
    pivot_df.rename(columns = {'NXT2-1':'NXT7'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'NXT7', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_TL4C.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_TL4C.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################

# Change directory Line-1 4C CycleTime Bottom
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Get current directory
current_directory = os.getcwd()

# Define the file path
file_path = 'CycleTime_BL4C.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    dct_H1 = pd.read_csv(file_path)
    
    # Convert ModuleName to string and concatenate with ModuleNo
    dct_H1["ModuleName"] = dct_H1['ModuleName'].astype(str) + "-" + dct_H1['ModuleNo'].astype(str)
    
    # Drop unnecessary columns
    dct_H1 = dct_H1.drop(['ModuleNo', 'GntaryNo', 'Placing Time', 'Average'], axis=1)
    
    # Pivot the DataFrame
    pivot_df = pd.pivot_table(dct_H1, values='CycleTime', index='LineName', columns='ModuleName', aggfunc='sum')
    
    # Add additional columns to the pivot table
    pivot_df['Lane'] = dct_H1['Lane'].iloc[0]  # Assuming all 'Lane' values are the same for the same 'LineName'
    pivot_df['Qty'] = dct_H1.groupby('LineName')['Qty'].sum()
    
    # Reset index
    pivot_df = pivot_df.reset_index()

    pivot_df.rename(columns = {'NXT1-1':'NXT1'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-2':'NXT2'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-3':'NXT3'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-4':'NXT4'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-5':'NXT5'}, inplace = True)
    pivot_df.rename(columns = {'NXT1-6':'NXT6'}, inplace = True)
    pivot_df.rename(columns = {'NXT2-1':'NXT7'}, inplace = True)

    # Rearrange columns in specified order
    pivot_df = pivot_df[['LineName', 'NXT1', 'NXT2', 'NXT3', 'NXT4', 'NXT5', 'NXT6', 'NXT7', 'Lane', 'Qty']]
    
    # Display the pivot DataFrame
    print('\n')
    print("Pivot Top", pivot_df)
    
    # Save the final DataFrame to a CSV file
    pivot_df.to_csv('CycleTime_BL4C.csv', index=False)
    
    # Print a message indicating the file has been saved
    print('\n')
    print("Final cycle time data saved to CycleTime_BL4C.csv")
else:
    print('The file does not exist.')

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print('\n')
print("\033[32;4m*******CYCLETIME DATA INSERTING*******\033[0m")

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')
Chd = os.getcwd()

######################################################################################################
#CT LINE1 TOP
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_TL1.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_tl1_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_tl1_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_tl1_data.insert(1, "Panel/Board", "0")
    cycle_time_tl1_data.insert(2, "Screen Printer", "30")
    cycle_time_tl1_data.insert(3, "SPI", "25")
    cycle_time_tl1_data.insert(4, "Reflow", "400")
    cycle_time_tl1_data.insert(5, "AOI", "25")
    cycle_time_tl1_data.insert(6, "Bottleneck", "")
    cycle_time_tl1_data.insert(7, "Cycle time", "")
    cycle_time_tl1_data.insert(8, "UPH", "")
    cycle_time_tl1_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Top_Line_Data.csv
    Top_line_data = pd.read_csv('Top_Line_Data.csv')
    
    # Append the top line data to the CycleTime_TL1 data
    merged_data = pd.concat([cycle_time_tl1_data, Top_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','AIMEXIII_1','AIMEXIII_2','AIMEXIII_3','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_TL1.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Top_Line_Data.csv moved into CycleTime_TL1.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE1 BOTTOM
######################################################################################################
    
# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_BL1.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_bl1_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_bl1_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_bl1_data.insert(1, "Panel/Board", "0")
    cycle_time_bl1_data.insert(2, "Screen Printer", "30")
    cycle_time_bl1_data.insert(3, "SPI", "25")
    cycle_time_bl1_data.insert(4, "Reflow", "400")
    cycle_time_bl1_data.insert(5, "AOI", "25")
    cycle_time_bl1_data.insert(6, "Bottleneck", "")
    cycle_time_bl1_data.insert(7, "Cycle time", "")
    cycle_time_bl1_data.insert(8, "UPH", "")
    cycle_time_bl1_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Bottom_Line_Data.csv
    Bottom_line_data = pd.read_csv('Bottom_Line_Data.csv')
    
    # Append the top line data to the CycleTime_BL1 data
    merged_data = pd.concat([cycle_time_bl1_data, Bottom_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','AIMEXIII_1','AIMEXIII_2','AIMEXIII_3','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_BL1.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Bottom_Line_Data.csv moved into CycleTime_BL1.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE2 TOP
######################################################################################################
    
# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_TL2.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_tl2_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_tl2_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_tl2_data.insert(1, "Panel/Board", "0")
    cycle_time_tl2_data.insert(2, "Screen Printer", "30")
    cycle_time_tl2_data.insert(3, "SPI", "25")
    cycle_time_tl2_data.insert(4, "Reflow", "400")
    cycle_time_tl2_data.insert(5, "AOI", "25")
    cycle_time_tl2_data.insert(6, "Bottleneck", "")
    cycle_time_tl2_data.insert(7, "Cycle time", "")
    cycle_time_tl2_data.insert(8, "UPH", "")
    cycle_time_tl2_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Top_Line_Data.csv
    Top_line_data = pd.read_csv('Top_Line_Data.csv')
    
    # Append the top line data to the CycleTime_TL2 data
    merged_data = pd.concat([cycle_time_tl2_data, Top_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','AIMEX-IIIC_1','AIMEX-IIIC_2','AIMEX-IIIC_3','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_TL2.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Top_Line_Data.csv moved into CycleTime_TL2.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE2 BOTTOM
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_BL2.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_bl2_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_bl2_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_bl2_data.insert(1, "Panel/Board", "0")
    cycle_time_bl2_data.insert(2, "Screen Printer", "30")
    cycle_time_bl2_data.insert(3, "SPI", "25")
    cycle_time_bl2_data.insert(4, "Reflow", "400")
    cycle_time_bl2_data.insert(5, "AOI", "25")
    cycle_time_bl2_data.insert(6, "Bottleneck", "")
    cycle_time_bl2_data.insert(7, "Cycle time", "")
    cycle_time_bl2_data.insert(8, "UPH", "")
    cycle_time_bl2_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Bottom_Line_Data.csv
    Bottom_line_data = pd.read_csv('Bottom_Line_Data.csv')
    
    # Append the top line data to the CycleTime_BL2 data
    merged_data = pd.concat([cycle_time_bl2_data, Bottom_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','AIMEX-IIIC_1','AIMEX-IIIC_2','AIMEX-IIIC_3','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_BL2.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Bottom_Line_Data.csv moved into CycleTime_BL2.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE3 TOP
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_TL3.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_tl3_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_tl3_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_tl3_data.insert(1, "Panel/Board", "0")
    cycle_time_tl3_data.insert(2, "Screen Printer", "30")
    cycle_time_tl3_data.insert(3, "SPI", "25")
    cycle_time_tl3_data.insert(4, "Reflow", "400")
    cycle_time_tl3_data.insert(5, "AOI", "25")
    cycle_time_tl3_data.insert(6, "Bottleneck", "")
    cycle_time_tl3_data.insert(7, "Cycle time", "")
    cycle_time_tl3_data.insert(8, "UPH", "")
    cycle_time_tl3_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Top_Line_Data.csv
    Top_line_data = pd.read_csv('Top_Line_Data.csv')
    
    # Append the top line data to the CycleTime_TL3 data
    merged_data = pd.concat([cycle_time_tl3_data, Top_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_TL3.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Top_Line_Data.csv moved into CycleTime_TL3.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE3 BOTTOM
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_BL3.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_bl3_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_bl3_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_bl3_data.insert(1, "Panel/Board", "0")
    cycle_time_bl3_data.insert(2, "Screen Printer", "30")
    cycle_time_bl3_data.insert(3, "SPI", "25")
    cycle_time_bl3_data.insert(4, "Reflow", "400")
    cycle_time_bl3_data.insert(5, "AOI", "25")
    cycle_time_bl3_data.insert(6, "Bottleneck", "")
    cycle_time_bl3_data.insert(7, "Cycle time", "")
    cycle_time_bl3_data.insert(8, "UPH", "")
    cycle_time_bl3_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Bottom_Line_Data.csv
    Bottom_line_data = pd.read_csv('Bottom_Line_Data.csv')
    
    # Append the top line data to the CycleTime_BL3 data
    merged_data = pd.concat([cycle_time_bl3_data, Bottom_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_BL3.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Bottom_Line_Data.csv moved into CycleTime_BL3.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE5 TOP
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_TL5.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_tl5_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_tl5_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_tl5_data.insert(1, "Panel/Board", "0")
    cycle_time_tl5_data.insert(2, "Screen Printer", "30")
    cycle_time_tl5_data.insert(3, "SPI", "25")
    cycle_time_tl5_data.insert(4, "Reflow", "400")
    cycle_time_tl5_data.insert(5, "AOI", "25")
    cycle_time_tl5_data.insert(6, "Bottleneck", "")
    cycle_time_tl5_data.insert(7, "Cycle time", "")
    cycle_time_tl5_data.insert(8, "UPH", "")
    cycle_time_tl5_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Top_Line_Data.csv
    Top_line_data = pd.read_csv('Top_Line_Data.csv')
    
    # Append the top line data to the CycleTime_TL5 data
    merged_data = pd.concat([cycle_time_tl5_data, Top_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_TL5.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Top_Line_Data.csv moved into CycleTime_TL5.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE5 BOTTOM
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_BL5.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_bl5_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_bl5_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_bl5_data.insert(1, "Panel/Board", "0")
    cycle_time_bl5_data.insert(2, "Screen Printer", "30")
    cycle_time_bl5_data.insert(3, "SPI", "25")
    cycle_time_bl5_data.insert(4, "Reflow", "400")
    cycle_time_bl5_data.insert(5, "AOI", "25")
    cycle_time_bl5_data.insert(6, "Bottleneck", "")
    cycle_time_bl5_data.insert(7, "Cycle time", "")
    cycle_time_bl5_data.insert(8, "UPH", "")
    cycle_time_bl5_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Bottom_Line_Data.csv
    Bottom_line_data = pd.read_csv('Bottom_Line_Data.csv')
    
    # Append the top line data to the CycleTime_BL5 data
    merged_data = pd.concat([cycle_time_bl5_data, Bottom_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_BL5.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Bottom_Line_Data.csv moved into CycleTime_BL5.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE-1 C4 TOP
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_TL4C.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_tl4C_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_tl4C_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_tl4C_data.insert(1, "Panel/Board", "0")
    cycle_time_tl4C_data.insert(2, "Screen Printer", "30")
    cycle_time_tl4C_data.insert(3, "SPI", "25")
    cycle_time_tl4C_data.insert(4, "Reflow", "400")
    cycle_time_tl4C_data.insert(5, "AOI", "25")
    cycle_time_tl4C_data.insert(6, "Bottleneck", "")
    cycle_time_tl4C_data.insert(7, "Cycle time", "")
    cycle_time_tl4C_data.insert(8, "UPH", "")
    cycle_time_tl4C_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Top_Line_Data.csv
    Top_line_data = pd.read_csv('Top_Line_Data.csv')
    
    # Append the top line data to the CycleTime_TL4C data
    merged_data = pd.concat([cycle_time_tl4C_data, Top_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_TL4C.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Top_Line_Data.csv moved into CycleTime_TL4C.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
#CT LINE-1 C4 BOTTOM
######################################################################################################

# Change directory to the location of the CSV files
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')

# Define the file path
file_path = 'CycleTime_BL4C.csv'

# Check if the file exists
if os.path.isfile(file_path):
    # Read the CSV file
    cycle_time_bl4C_data = pd.read_csv(file_path)
    
    # Get current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d %I:%M:%S %p")
    
    # Add current date and time to the header
    cycle_time_bl4C_data.insert(0, "Current_DateTime", current_datetime)
    cycle_time_bl4C_data.insert(1, "Panel/Board", "0")
    cycle_time_bl4C_data.insert(2, "Screen Printer", "30")
    cycle_time_bl4C_data.insert(3, "SPI", "25")
    cycle_time_bl4C_data.insert(4, "Reflow", "400")
    cycle_time_bl4C_data.insert(5, "AOI", "25")
    cycle_time_bl4C_data.insert(6, "Bottleneck", "")
    cycle_time_bl4C_data.insert(7, "Cycle time", "")
    cycle_time_bl4C_data.insert(8, "UPH", "")
    cycle_time_bl4C_data.insert(9, "Cycle Remarks", "")
    
    # Read the data from Bottom_Line_Data.csv
    Bottom_line_data = pd.read_csv('Bottom_Line_Data.csv')
    
    # Append the top line data to the CycleTime_BL4C data
    merged_data = pd.concat([cycle_time_bl4C_data, Bottom_line_data], axis=1)

    merged_data.rename(columns = {'Qty':'Total Component'}, inplace = True)
    merged_data = merged_data[['LineName','Identifier','Version','Position','Current_DateTime','Production Stage','Lane','Panel/Board','Total Component','Screen Printer','SPI','NXT1','NXT2','NXT3','NXT4','NXT5','NXT6','Reflow','AOI','Setup Details','Bottleneck','Cycle time','UPH','Cycle Remarks']]
    
    # Save the merged data to CycleTime_BL4C.csv
    merged_data.to_csv(file_path, index=False)
    
    print('\n')
    print("Data from Bottom_Line_Data.csv moved into CycleTime_BL4C.csv")
else:
    # If the file does not exist, print a message
    print('The file does not exist.')

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print('\n')
print("\033[32;4m*******CYCLETIME DATA TRANSFERING*******\033[0m")
######################################################################################################


# Change directory to the CycleTime folder
os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime')
print('\n')
print('CycleTime---')
print('\n')
# Function to append data to existing Excel sheet or create a new one
def append_to_excel(csv_file, excel_file):
    # Read the CSV file
    df = pd.read_csv(csv_file)
    
    # Get the sheet name from the first row of the CSV file
    line_name = df.iloc[0]['LineName']
    
    # Check if the Excel file exists
    if os.path.exists(excel_file):
        # Load existing Excel file
        existing_sheets = pd.read_excel(excel_file, sheet_name=None)
        
        # Check if the sheet exists in the Excel file
        if line_name in existing_sheets:
            # Append data to existing sheet
            existing_df = existing_sheets[line_name]
            combined_data = pd.concat([existing_df, df], ignore_index=True)
            existing_sheets[line_name] = combined_data
        else:
            # Create a new sheet and write data
            existing_sheets[line_name] = df
    
    else:
        # If the Excel file doesn't exist, create a new one and write data
        existing_sheets = {line_name: df}
    
    # Write all sheets back to Excel file
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for sheet_name, sheet_data in existing_sheets.items():
            sheet_data.to_excel(writer, sheet_name=sheet_name, index=False)

# Define the Excel file
excel_file = 'CycleTime.xlsx'

# List of CSV files
csv_files = ['CycleTime_BL1.csv', 'CycleTime_TL1.csv', 'CycleTime_BL2.csv', 'CycleTime_TL2.csv', 'CycleTime_BL3.csv', 'CycleTime_TL3.csv', 'CycleTime_BL5.csv', 'CycleTime_TL5.csv', 'CycleTime_BL4C.csv', 'CycleTime_TL4C.csv']

# Filter out the CSV files that exist in the CycleTime folder
existing_csv_files = [os.path.join('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime', file) for file in csv_files if os.path.isfile(os.path.join('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/CycleTime', file))]

# Upload data to Excel file for each existing CSV file
os.chdir('D:/NX_BACKWORK/Database_File/SMT_CT')
for csv_file in existing_csv_files:
    append_to_excel(csv_file, excel_file)

######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
######################################################################################################
print("\033[32;4m*******FeederSetup Input*******\033[0m")
######################################################################################################
# Input values for cell B3 and Rev A1
print('\n')
dL1 = input("\033[93mEnter BOM Name :\033[0m")
print('\n')
dL2 = input("\033[93mEnter Feeder Name :\033[0m")
value_B3 = dL2[:12]  # Take only the first 12 characters from dL2
print('\n')
Revision = input("\033[93mEnter the program for Revision: \033[0m")
print('\n')
dLine123 = input("\033[93mEnter the Line: \033[0m")

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
Chd = os.getcwd()

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Upload"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)
    
yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\AVL & Polarity Check"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)
    
yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\PartNumber"

if not os.path.isdir(yourfolder):
    print('Folder Not Exist')
    os.makedirs(yourfolder)

os.getcwd()
#Chd= os.chdir('D:\\NX_BACKWORK')
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

if os.path.exists("Feeder_List_OPT.xlsx"):
    os.remove("Feeder_List_OPT.xlsx")
else:
    print("The file does not exist")

if os.path.exists("Feeder_List_OPB.xlsx"):
    os.remove("Feeder_List_OPB.xlsx")
else:
    print("The file does not exist")

if os.path.exists("Upload-Data.xlsx"):
    os.remove("Upload-Data.xlsx")
else:
    print("The file does not exist")

os.getcwd()
#Chd= os.chdir('D:\\NX_BACKWORK')
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')
Chd = os.getcwd()

##########################################################################################################################################

#bil2 = pyfiglet.figlet_format("FeederSetup Progress", width = 150)
print('\n')
print('\033[92;4m******FeederSetup Progress******\033[0m')
print('\n')
##########################################################################################################################################

#LINE1T

file_path = 'FeederSetup_TL1.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)
        
except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL1.csv', encoding="utf-8",index_col=False, skiprows=range(2))
    
    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 400:
        # Continue with the rest of your code
        print(f"dt_H1 line count: {len(dt_H1)}")
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 403.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in TOP Feeder '403'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_TL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in TOP Feeder '401'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE TOP NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)
    
    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_TL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_TL1.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_TL1: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')
    
    df1["ModelName"] = df1['ModelName'].astype(str) +"-"+ df1['ModuleNumber']

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('1-3-2-','3-2-')

    df1['Location'] = df1['Location'].str.replace('1-3-1-','3-1-')

    df1['Location'] = df1['Location'].str.replace('1-2-2-','2-2-')

    df1['Location'] = df1['Location'].str.replace('1-2-1-','2-1-')

    df1['Location'] = df1['Location'].str.replace('1-1-2-','1-2-')

    df1['Location'] = df1['Location'].str.replace('1-1-1-','1-1-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1B

file_path = 'FeederSetup_BL1.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL1.csv', encoding="utf-8",index_col=False, skiprows=range(2))
    
    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 400:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 403.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in BOT Feeder '403'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_BL1.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

            # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-1 Slot Count in BOT Feeder '401'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE BOT NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_BL1.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_BL1.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_BL1: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code


#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')
    
    df1["ModelName"] = df1['ModelName'].astype(str) +"-"+ df1['ModuleNumber']

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('1-3-2-','3-2-')

    df1['Location'] = df1['Location'].str.replace('1-3-1-','3-1-')

    df1['Location'] = df1['Location'].str.replace('1-2-2-','2-2-')

    df1['Location'] = df1['Location'].str.replace('1-2-1-','2-1-')

    df1['Location'] = df1['Location'].str.replace('1-1-2-','1-2-')

    df1['Location'] = df1['Location'].str.replace('1-1-1-','1-1-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE2T

file_path = 'FeederSetup_TL2.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL2.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 400:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 403.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in TOP Feeder '403'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_TL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in TOP Feeder '401'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE TOP NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_TL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_TL2.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_TL2: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1','3')

    df1['Location'] = df1['Location'].str.replace('2-1','2')

    df1['Location'] = df1['Location'].str.replace('1-1','1')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE2B

file_path = 'FeederSetup_BL2.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL2.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 400:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 403.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in BOT Feeder '403'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_BL2.csv', encoding="utf-8",index_col=False, skiprows=range(2, 401), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-2 Slot Count in BOT Feeder '401'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
    
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE BOT NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_BL2.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_BL2.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_BL2: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    #df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('3-1','3')

    df1['Location'] = df1['Location'].str.replace('2-1','2')

    df1['Location'] = df1['Location'].str.replace('1-1','1')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
        #df1.to_excel(writer, sheet_name="FS_upload", index=False)
        #df2.to_excel(writer, sheet_name="FS_Count", index=False)
        #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
        #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)   
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE3T

file_path = 'FeederSetup_TL3.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL3.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 170:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 173.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in TOP Feeder '173'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_TL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in TOP Feeder '171'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE TOP NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_TL3: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE3B

file_path = 'FeederSetup_BL3.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL3.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 170:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 173.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in BOT Feeder '173'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_BL3.csv', encoding="utf-8",index_col=False, skiprows=range(2, 171), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-3 Slot Count in BOT Feeder '171'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
    
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE BOT NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_BL3: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE5T

file_path = 'FeederSetup_TL5.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL5.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 182:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 185.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-5 Slot Count in TOP Feeder '185'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_TL5.csv', encoding="utf-8",index_col=False, skiprows=range(2, 183), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-5 Slot Count in TOP Feeder '183'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE TOP NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_TL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_TL5.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_TL5: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE5B

file_path = 'FeederSetup_BL5.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL5.csv', encoding="utf-8",index_col=False, skiprows=range(2))

        # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 182:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 185.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-5 Slot Count in BOT Feeder '185'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_BL5.csv', encoding="utf-8",index_col=False, skiprows=range(2, 183), nrows=3)

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-5 Slot Count in BOT Feeder '185'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
    
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE BOT NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_BL3.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_BL5.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_BL5: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)
    
    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1T-C4

file_path = 'FeederSetup_TL4C.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_TL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 227:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 230.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in TOP Feeder '230'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_TL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in TOP Feeder '228'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]

    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

    #dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE TOP NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_TL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_TL4C.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_TL4C: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')

#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''
    
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb1 = load_workbook(Feeder_List_OPT)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPT ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPT.xlsx"
    wb2 = load_workbook(Feeder_List_OPT)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
        #copy from wb1
        c = ws1.cell(row=row, column=10)
        #paste in ws2
        ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPT.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPT.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "TOP"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['TOP_Side'] = "TOP"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "TOP"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "TOP"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)    
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)
pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#LINE1B-C4

file_path = 'FeederSetup_BL4C.csv'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    dt_H1 = pd.read_csv('FeederSetup_BL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2))

    # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 227:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        print("dt_H1 is either None or its length is not equal to 230.")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in BOT Feeder '230'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    dt_H1 = pd.read_csv('FeederSetup_BL4C.csv', encoding="utf-8",index_col=False, skiprows=range(2, 228), nrows=3)

        # Check if dt_H1 is defined and the line count is 351
    if dt_H1 is not None and len(dt_H1) == 3:
        # Continue with the rest of your code
        print(dt_H1)
    else:
        # Show error message
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = "Error: Either Check the FeederSetup.csv Line-4C Slot Count in BOT Feeder '228'."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
    
    dt_H1['TotalSlots'] = ''
    dt_H1['TotalSlots'].loc[0] = dt_H1['JobFolder'].loc[2]
    
    dt_H1['PlacedParts'] = ''
    dt_H1['PlacedParts'].loc[0] = dt_H1['JobName'].loc[2]

#dt_H1.drop(dt_H1.iloc[:, 17:26], inplace=True, axis=1)

    dt_H1['Col1'] = dt_H1['JobName'].str[13:]

    dt_H1['Col2'] = dt_H1['TopBottom'].astype(str).str[:1]

    print(dt_H1.drop(index=[1, 2]))

    dt_H1 = dt_H1.drop(dt_H1.index[[1,2]])

    dt_H1['Side'] = dt_H1['Col1']+"-"+ dt_H1['Col2']

    dt_H1['Side'] = dt_H1['Side'].replace('T-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('B-0','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-0','TOP')

    dt_H1['Side'] = dt_H1['Side'].replace('T&B-1','BOT')

    dt_H1['Side'] = dt_H1['Side'].replace('B-1','BOT')
    
    # Check if Col1 contains only valid values
    
    valid_values = {'T', 'B', 'T&B'}
    invalid_rows = dt_H1[~dt_H1['Col1'].isin(valid_values)]

    # Handle invalid values with a pop-up
    if not invalid_rows.empty:
        # Initialize tkinter
        root = tk.Tk()
        root.withdraw()  # Hide the main tkinter window

        # Format the invalid rows as a string
        invalid_entries = invalid_rows.to_string(index=False)

        # Show the error message with details of invalid rows
        messagebox.showerror("Invalid Data", f"Error: Invalid values found in 'Col1' PROGRAM NAME SIDE BOT NOT MENTION.\n\nInvalid Entries:\n{invalid_entries}")

        # Exit after showing the error
        root.destroy()
    else:
        # Proceed with the rest of the logic
        print("Processing completed successfully.")
        print(dt_H1)

    dt_H1.drop(dt_H1.iloc[:, 13:17], inplace=True, axis=1)

    N9_Col = dt_H1.pop('Side') # col-10

    dt_H1.insert(9, 'Side', N9_Col)

    dt_H1.drop(dt_H1.iloc[:, 10:11], inplace=True, axis=1)

    dt_H1.drop(dt_H1.iloc[:, 15:17], inplace=True, axis=1)

    dt_H1.insert(15, 'CATEGORY', '')
    dt_H1.insert(16, 'MODEL NAME', '')
    dt_H1.insert(17, 'CURRENT REVISION', '')
    dt_H1.insert(18, 'MODIFIED  DATE', '')
    dt_H1.insert(19, 'MODIFICATION DESCRIPTION', '')
    dt_H1.insert(20, 'BOM ECO NUMBER', '')
    dt_H1['Verify-DateTime'] = datetime.now()
    dt_H1.rename(columns = {'JobFolder':'CUSTOMER NAME'}, inplace = True)
    dt_H1.rename(columns = {'JobName':'PROGRAM NAME'}, inplace = True)
    dt_H1.rename(columns = {'Revision':'PRO. Rev'}, inplace = True)
    dt_H1.rename(columns = {'ModifiedDate':'PRO.ModifiedDate'}, inplace = True)

    dt_H1 = dt_H1[['CUSTOMER NAME','PROGRAM NAME','PRO. Rev','PRO.ModifiedDate','Comments','Product','LogOnUser','Line','SetupName','Side','PanelLength','PanelWidth','PanelThickness','TotalSlots','PlacedParts','CATEGORY','MODEL NAME','CURRENT REVISION','MODIFIED  DATE','MODIFICATION DESCRIPTION','BOM ECO NUMBER','Verify-DateTime']]
    #CUSTOMER NAME	PROGRAM NAME	PRO. Rev	PRO.ModifiedDate	Comments	Product	LogOnUser	Line	SetupName	Side	PanelLength	PanelWidth	PanelThickness	TotalSlots	PlacedParts	CATEGORY	MODEL NAME	CURRENT REVISION	MODIFIED  DATE	MODIFICATION DESCRIPTION	BOM ECO NUMBER

    print(dt_H1)
#-----------------------------------------------------------------------------------------------------------------------#

#print(dt_H1.drop(index=[1, 2]))

# NOTE df = pd.read_csv(filename, skiprows=range(2, 20000), nrows=10000)
#df = pd.read_csv(csv_filepath , skiprows=2, encoding="utf-8",index_col=False)

#-----------------------------------------------------------------------------------------------------------------------#

    #df1 = pd.read_csv('FeederSetup_BL4C.csv', skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 

# Specify the columns you want to read
    columns_to_read = ['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList']

    try:
        df1 = pd.read_csv('FeederSetup_BL4C.csv', skiprows=2, usecols=columns_to_read, encoding="utf-8", index_col=False)

        # Check if all the specified columns are present in the DataFrame
        if all(column in df1.columns for column in columns_to_read):
            print("All columns are present in the DataFrame.")
        else:
            # Show error message if any columns are missing
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            missing_columns = [column for column in columns_to_read if column not in df1.columns]
            error_message = f"The following columns are missing in FeederSetup: {', '.join(missing_columns)}"
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

    except Exception as e:
            # Handle the exception gracefully
            error_message = f"An error occurred FeederSetup_BL4C: {e}"

            # Show error message in a pop-up box
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showerror("Error", error_message)
            sys.exit(1)  # Exit the program with an error code

#df1 = pd.read_csv(csv_filepath,skiprows=2 , usecols=['LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'], encoding="utf-8",index_col=False) 
# NOTE Line to usecol to call desire column only 'LineName', 'OrderNum', 'ModelName', 'ModuleNumber', 'SideNo', 'PartNumber', 'FeederName', 'Status', 'Location', 'PackageName', 'PartComment', 'PMABAR', 'ChuteType', 'FeederType', 'TapeWidth', 'FeedPitch', 'PTPMNH', 'QTY', 'RefList'

    df1.dropna(subset=['RefList'], inplace=True)

    df1['SideNo'] = df1['SideNo'].astype(str).str.replace('.', '')

    df1['SideNo'] = df1['SideNo'].str.replace('10', '2')

    df1['SideNo'] = df1['SideNo'].str.replace('00', '1')

    print (df1)

    df1['ModuleNumber'] = df1['ModuleNumber'].astype(str).str.replace('.0', '')

    #df1['ModelName'] = df1['ModelName'].str.replace('1','')

    df1.rename(columns = {'Location':'Lock'}, inplace = True)

    df1.rename(columns = {'LineName':'Location'}, inplace = True)

    df1['OrderNum'] = df1['OrderNum'].str.replace('1','0')

    df1["Location"] = df1['OrderNum'].astype(str) +"-"+ df1['ModuleNumber'].astype(str) +"-"+ df1['SideNo'].astype(str) +"-"+ df1["Lock"]

    df1['Location'] = df1['Location'].str.replace('2-1-1-','7-')

    df1['Location'] = df1['Location'].str.replace('0-1-1-','1-')

    df1['Location'] = df1['Location'].str.replace('0-2-1-','2-')

    df1['Location'] = df1['Location'].str.replace('0-3-1-','3-')

    df1['Location'] = df1['Location'].str.replace('0-4-1-','4-')

    df1['Location'] = df1['Location'].str.replace('0-5-1-','5-')

    df1['Location'] = df1['Location'].str.replace('0-6-1-','6-')

    F1_col = df1.pop('PartNumber') # col-1

    df1.insert(1, 'PartNumber', F1_col)

    df1.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)

    S2_col = df1.pop('FeederName') # col-2

    df1.insert(2, 'FeederName', S2_col)

    T3_col = df1.pop('FeederType') # col-3 rename to type

    df1.insert(3, 'FeederType', T3_col)

    df1.rename(columns = {'FeederType':'Type'}, inplace = True)

    F4_col = df1.pop('TapeWidth') # col-4 rename to size

    df1.insert(4, 'TapeWidth', F4_col)

    df1.rename(columns = {'TapeWidth':'Size'}, inplace = True)

    F5_col = df1.pop('FeedPitch') # col-5 

    df1.insert(5, 'FeedPitch', F5_col)

    S6_col = df1.pop('PTPMNH') # col-6 rename to Part Height  

    df1.insert(6, 'PTPMNH', S6_col)

    df1.rename(columns = {'PTPMNH':'Part Height'}, inplace = True)

    S7_col = df1.pop('Status') # col-7

    df1.insert(7, 'Status', S7_col)

    E8_col = df1.pop('QTY') # col-8 

    df1.insert(8, 'QTY', E8_col)

    df1.drop(df1.iloc[:, 9:10], inplace=True, axis=1)

    df1.drop(df1.iloc[:, 10:17], inplace=True, axis=1)

    extracted_col = dt_H1["Side"] 

    df1.insert(9, "Side", extracted_col)

    #NEW PN# PARTNO
    df1['PartNO'] = "PN#"
    df1["F_Part_No"] = df1['PartNO'].astype(str) +""+ df1['F_Part_No'].astype(str)
    del df1['PartNO']
    df1['F_Part_No'] = df1['F_Part_No'].str.replace('.0','')
    
#dt_H1.drop(dt_H1.iloc[:, 13:15], inplace=True, axis=1)'''

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)
        
    # Function to keep only values starting with '1:' A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z PANELUP CODE
    def keep_first_sequence(text):
        if pd.isna(text):  # Handle missing values
            return ""
        pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[a-zA-Z0-9._-]+' #pattern = r'1:[aA-zZ]+\d*\-*\.*\_*\d*' #pattern = r'1:[A-Z]\d+'  # Match only sequences starting with "1:" 
        matches = re.findall(pattern, str(text))  # Convert to string and find matches
        return " ".join(matches)  # Join them back into a string
    
    # Load the Excel file
    file_path = "D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"  # Change this to your actual file path
    dt_H1 = pd.read_excel(file_path, sheet_name="S1")
    df1 = pd.read_excel(file_path, sheet_name="FeederSetup1")

    # Apply the function to the 'RefList' column
    df1['RefList'] = df1['RefList'].apply(keep_first_sequence)

    # Save the modified DataFrame back to a new Excel file
    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", mode='w') as writer:  #"C:/Users/Bala Ganesh/Documents/Python/filename_OP.xlsx"

        dt_H1.to_excel(writer, sheet_name="S1", index=False)
        df1.to_excel(writer, sheet_name="FeederSetup1", index=False)

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb1 = load_workbook(Feeder_List_OPB)
    ws1 = wb1.active
    ws1 = wb1.worksheets[0]

    Feeder_List_OPB ="D:\\NX_BACKWORK\\Feeder Setup_PROCESS\\#Output\\FeederSetup\\Feeder_List_OPB.xlsx"
    wb2 = load_workbook(Feeder_List_OPB)
    ws2 = wb2.active
    ws2 = wb2.worksheets[1]

    print(ws1)

    print(ws2)

    for row in range(1, 10):
    #copy from wb1
            c = ws1.cell(row=row, column=10)
    #paste in ws2
            ws2.cell(row=row-0, column=10, value=c.value)

    print(ws2)

    wb2.save(str('Feeder_List_OPB.xlsx'))

    df1 = pd.read_excel('Feeder_List_OPB.xlsx','FeederSetup1', index_col=False)

    df1['Side'] = df1['Side'].fillna(method='ffill') # NOTE forword fukk added

    df1.rename(columns = {'RefList':'F_Ref_List'}, inplace = True)

    df2 = df1['F_Ref_List'].str.split(' ',expand=True) # RL1 = df to split the reflit column only .str.split ' space ' expand true it will expand no of space

    df3 = pd.concat([df1, df2], axis=1) # NOTE 'reflist next create column 0,1,2,3,4,5,6,7to Nth digite' NOTE df3 = "it split the 0-N value adding brfore ref'. if ihad df2 = "page contain only 0 to N value" 

# NOTE reflist column to next axis column this line df = nothing df1 split colum next to

#print("Column headers from list(df.columns.values):", list(df2.columns.values)) # NOTE this line for record to seem the no of header value (df.Columns.Values) no of row count formed. 

    df2 = df3.melt(id_vars=['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List'], var_name='df.columns.values')

    df2.explode ('F_Ref_List')

    df2.dropna(subset=['value'], inplace=True) # NOTE subset the value column 

    df2.drop(df2.iloc[:, 5:9], inplace=True, axis=1) # NOTE Remove the [FeedPitch,Part Height,Status,QTY]

    df2.drop(df2.iloc[:, 7:9], inplace=True, axis=1)

    df2.rename(columns = {'value':'F_Ref_List'}, inplace = True)

    df2['F_Ref_List'] = df2['F_Ref_List'].str.replace('1:','')

#df3 = df3.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'PartHeight', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

# NOTE df3.loc [variable] delete all col after reflist 

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace(' ',',')

    df3['F_Ref_List'] = df3['F_Ref_List'].str.replace('1:','')

    df3_1 = df3.explode('F_Ref_List') # NOTE line to create dummy new page with old content

#print (df3_1) # NOTE line to create dummy new page with old content

    df3_1.insert(12,'RefList1','') # NOTE line to create dummy new page with old content

    df3_1.drop(df3_1.iloc[:, 12:13], inplace=True, axis=1)

    df3_1.insert(9, 'Tray Dir','')
    df3_1.insert(10, 'PartComment','')
    df3_1.insert(11, 'Barcode Label','')

    df4_1 = df3_1['Size'].value_counts()
    df4_1['Feedersize'] = "BOT"
    df5_1 = df3_1['Side'].value_counts()
    df5_1['BOT_Side'] = "BOT"
    df6_1 = df3_1['FeederName'].value_counts()
    df6_1['FeederSize'] = "BOT"
    df7_1 = df3_1['Type'].value_counts()
    df7_1['FeederType'] = "BOT"

    df3['F_Ref_List'] = df3['F_Ref_List'] .str.strip('[]').str.split(',')

    df3.to_dict()

    df3.explode ('F_Ref_List',ignore_index=True)

    df4 = df3.explode('F_Ref_List',ignore_index=True)

    df4 = df4.loc[:,['Location', 'F_Part_No', 'FeederName', 'Type', 'Size', 'FeedPitch', 'Part Height', 'Status', 'QTY','Side', 'ModelName', 'F_Ref_List']]

    df4.head()

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx") as writer:

        dt_H1.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df.to_excel(writer, sheet_name="FeederSetup0", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
        df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)   
        df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        df4_1.to_excel(writer, sheet_name="Size", index=TRUE)
        df5_1.to_excel(writer, sheet_name="Side", index=TRUE)
        df6_1.to_excel(writer, sheet_name="FeederName", index=TRUE)
        df7_1.to_excel(writer, sheet_name="Type", index=TRUE)

pass
print('The file does not exist.')

##########################################################################################################################################

##########################################################################################################################################

#bil3 = pyfiglet.figlet_format("FeederSetup Progress Merge", width = 150)
print('\n')
print("\033[92;4m******FeederSetup Progress Merge******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################

#LEN MERGE TOP AND BOT

data_file_folder = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup'

df=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Home'))
df1=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df1.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Upload'))
df2=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df2.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FL_Verify'))
df3=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df3.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Size'))
df4=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df4.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Side'))
df5=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df5.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='FeederName'))

df6=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        df6.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Type'))

len(df)
df_master1 = pd.concat(df, axis=0)
len(df1)
df_master2 = pd.concat(df1, axis=0)
len(df2)
df_master3 = pd.concat(df2, axis=0)
len(df3)
df_master4 = pd.concat(df3, axis=0)
len(df4)
df_master5 = pd.concat(df4, axis=0)
len(df5)
df_master6 = pd.concat(df5, axis=0)
len(df6)
df_master7 = pd.concat(df6, axis=0)

    #NEW PN# PARTNO
    #df3_1['F_Part_No'] = df3_1['F_Part_No'].str.replace('PN#','')
    #df4['F_Part_No'] = df4['F_Part_No'].str.replace('PN#','')

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx") as writer:
    #df_master.to_excel('masterfile.xlsx',index=False)
        df_master1.to_excel(writer, sheet_name="Home", index=False)
        df_master2.to_excel(writer, sheet_name="FeederSetup", index=False)
        df_master3.to_excel(writer, sheet_name="FeederCol", index=False)
        df_master4.to_excel(writer, sheet_name="FeederSize", index=False)
        df_master5.to_excel(writer, sheet_name="Total side Count", index=False)
        df_master6.to_excel(writer, sheet_name="FeederName", index=False)
        df_master7.to_excel(writer, sheet_name="Type", index=False)

Chd = os.getcwd()
Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')

if os.path.exists("Feeder_List_OPT.xlsx"):
    os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPT.xlsx", "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPT.xlsx")
else:
    print("The file does not exist")

Chd = os.getcwd()
Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup')

if os.path.exists("Feeder_List_OPB.xlsx"):
    os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/FeederSetup/Feeder_List_OPB.xlsx", "D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Feeder_List_OPB.xlsx")
else:
    print("The file does not exist")

Chd = os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
Chd = os.getcwd()

# Process Feeder_List_OPT.xlsx
try:
    xls = pd.ExcelFile('Feeder_List_OPT.xlsx', engine='openpyxl')
    dffst11 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Home")
    dffst12 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FL_Upload")
    dffst13 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FL_Verify")
    dffst14 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Size")
    dffst15 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Side")
    dffst16 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="FeederName")
    dffst17 = pd.read_excel("Feeder_List_OPT.xlsx", sheet_name="Type")

    dffst12['F_Part_No'] = dffst12['F_Part_No'].str.replace('PN#', '')
    dffst13['F_Part_No'] = dffst13['F_Part_No'].str.replace('PN#', '')

    with pd.ExcelWriter('Feeder_List_OPT.xlsx', engine='openpyxl', mode='w') as writer:
        dffst11.to_excel(writer, sheet_name="Home", index=False)
        dffst12.to_excel(writer, sheet_name="FL_Upload", index=False)
        dffst13.to_excel(writer, sheet_name="FL_Verify", index=False)
        dffst14.to_excel(writer, sheet_name="Size", index=False)
        dffst15.to_excel(writer, sheet_name="Side", index=False)
        dffst16.to_excel(writer, sheet_name="FeederName", index=False)
        dffst17.to_excel(writer, sheet_name="Type", index=False)

except FileNotFoundError:
    print("Feeder_List_OPT.xlsx not found. Skipping processing.")

# Process Feeder_List_OPB.xlsx
try:
    xls = pd.ExcelFile('Feeder_List_OPB.xlsx', engine='openpyxl')
    dffsb11 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Home")
    dffsb12 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FL_Upload")
    dffsb13 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FL_Verify")
    dffsb14 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Size")
    dffsb15 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Side")
    dffsb16 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="FeederName")
    dffsb17 = pd.read_excel("Feeder_List_OPB.xlsx", sheet_name="Type")

    dffsb12['F_Part_No'] = dffsb12['F_Part_No'].str.replace('PN#', '')
    dffsb13['F_Part_No'] = dffsb13['F_Part_No'].str.replace('PN#', '')

    with pd.ExcelWriter('Feeder_List_OPB.xlsx', engine='openpyxl', mode='w') as writer:
        dffsb11.to_excel(writer, sheet_name="Home", index=False)
        dffsb12.to_excel(writer, sheet_name="FL_Upload", index=False)
        dffsb13.to_excel(writer, sheet_name="FL_Verify", index=False)
        dffsb14.to_excel(writer, sheet_name="Size", index=False)
        dffsb15.to_excel(writer, sheet_name="Side", index=False)
        dffsb16.to_excel(writer, sheet_name="FeederName", index=False)
        dffsb17.to_excel(writer, sheet_name="Type", index=False)
except FileNotFoundError:
    print("Feeder_List_OPB.xlsx not found. Skipping processing.")

##########################################################################################################################################

##########################################################################################################################################

#bil4 = pyfiglet.figlet_format("Program Master List", width = 100)
print('\n')
print("\033[92;4m******Program Master List******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################

#Master list program update and LOG

os.getcwd()
Chd= os.chdir("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified")
Chd = os.getcwd()
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
dfpm1 = pd.read_excel('FeederSetup.xlsx', sheet_name='Home')

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Pro_ML.xlsx") as writer:
    #df_master.to_excel('masterfile.xlsx',index=False)
        dfpm1.to_excel(writer, sheet_name="Home", index=False)

if os.path.exists("Pro_ML.xlsx"):
    os.rename("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/Pro_ML.xlsx" , "D:/NX_BACKWORK/Database_File/SMT_Master_List/Pro_ML.xlsx")
else:
    print("The file does not exist")

data_file_folder = 'D:/NX_BACKWORK/Database_File/SMT_Master_List'

dspm1=[]
for file in os.listdir(data_file_folder):
    if file.endswith('.xlsx'):
        print('Loading file {0}...'.format(file))
        dspm1.append(pd.read_excel(os.path.join(data_file_folder,file),sheet_name='Home'))

len(dspm1)
dsf_master1 = pd.concat(dspm1, axis=0)

with pd.ExcelWriter("D:/NX_BACKWORK/Database_File/SMT_Master_List/Program_Master_List.xlsx") as writer:
    dsf_master1.to_excel(writer, sheet_name="Home", index=False)

print("Tranfer Complete...")

os.getcwd()
Chd= os.chdir("D:/NX_BACKWORK/Database_File/SMT_Master_List")
Chd = os.getcwd()

print("Del Start....")

if os.path.exists("Pro_ML.xlsx"):
    os.remove("Pro_ML.xlsx")
else:
    print("The file does not exist")

##########################################################################################################################################

##########################################################################################################################################

#bil5 = pyfiglet.figlet_format("BOM Manipulation", width = 100)
print('\n')
print("\033[92;4m******BOM Manipulation******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################

#BOM MANIPULATE

try:
    # BOM MANIPULATION
    os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    file_path = 'BOM.xlsx'

    if os.path.isfile(file_path):
        ds1 = pd.read_excel(file_path, index_col=False)
    else:
        # Try reading as '.xls' format if '.xlsx' fails
        file_path = 'BOM.xls'
        ds1 = pd.read_excel(file_path, index_col=False)

    dfbom1 = ds1

except ValueError:
    dfbom1 = pd.read_excel(file_path,index_col=False) 

except Exception as e:
    # Handle the exception gracefully
    error_message = f"An error occurred: {e}"

    # Show error message in a pop-up box
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showerror("Error", error_message)
    sys.exit(1)  # Exit the program with an error code

# Define your column lists
column_list_1 = ['Material', 'AltItemGroup', 'Priority', 'Long. Description', 'Ref.Designator/Circuit Reference', 'Quantity', 'Material Group']
column_list_2 = ['Internal P/N', 'Group', 'Priority', 'Description', 'Ref.Designator', 'Qty', 'SMT/THT/Mech']

# Check which column list is present in the DataFrame
if all(column in ds1.columns for column in column_list_1):
    columns_to_use = column_list_1
elif all(column in ds1.columns for column in column_list_2):
    columns_to_use = column_list_2
else:
    # Show error message if none of the column lists is present
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    missing_columns = [column for column_list in [column_list_1, column_list_2] for column in column_list if column not in ds1.columns]
    error_message = f"The following columns are missing: {', '.join(missing_columns)}"
    error_msgbm1 = f"The following columns are missing: in SAP BOM\n'Material'\n'AltItemGroup'\n'Priority'\n'Long. Description'\n'Ref.Designator/Circuit Reference'\n'Quantity'\n'Material Group'"
    error_msgbm2 = f"The following columns are missing: in Internal BOM\n'Internal P/N'\n'Group'\n'Priority'\n'Description'\n'Ref.Designator'\n'Qty'\n'SMT/THT/Mech'"
    messagebox.showerror("Error", error_message)
    messagebox.showerror("Error", error_msgbm1)
    messagebox.showerror("Error", error_msgbm2)
    sys.exit(1)  # Exit the program with an error code

# Continue with the rest of your code using 'columns_to_use'
print(f"Using columns: {columns_to_use}")

# Rest of your code here
# ...

ds1.rename(
    columns={'Material':"PartNumber", 'AltItemGroup':"Group", 'Priority':'Priority', 'Long. Description':'Long Des', 'Ref.Designator/Circuit Reference':'RefList', 'Quantity':'Qty','Material Group':'Shape'},
    inplace=True,
)

ds1.rename(
    columns={'Internal P/N':"PartNumber", 'Group':"Group", 'Priority':'Priority', 'Description':'Long Des', 'Ref.Designator':'RefList', 'Qty':'Qty','SMT/THT/Mech':'Shape'},
    inplace=True,
)

print(ds1)

ds2 = ds1[ds1['Priority'].isin([0, 1])]

# Assuming ds2 is your DataFrame and 'PartNumber' and 'RefList' are the columns you want to check
part_number_column = ds2['PartNumber']
ref_list_column = ds2['RefList']

# Flag to check if an empty value is found
empty_value_found = False

# Iterate through both columns simultaneously using iterrows
for index, (part_number_value, ref_list_value) in ds2[['PartNumber', 'RefList']].iterrows():
    # Check if the 'RefList' value is empty (NaN or None)
    if pd.isna(ref_list_value):
        print(f"Error: Empty value found in 'RefList' for 'PartNumber' {part_number_value}")
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = f"Empty value found in 'RefList' for 'PartNumber' {part_number_value}. Program will stop."
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
        #raise ValueError("Empty value found in 'RefList'")
        #empty_value_found = True
        #break  # Stop the iteration when the first empty value is found

# If no empty values are found, print the 'PartNumber' column
if not empty_value_found:
    print(part_number_column)
    # Continue with the rest of your program

#file_name ="output.xlsx"
#ds1.to_excel(file_name)

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM/BOM_List_OP.xlsx") as writer:
    ds1.to_excel(writer, sheet_name="Orginal_BOM", index=False)
    ds2.to_excel(writer, sheet_name="BOM", index=False)

    pass
    print('The file does not exist.')

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
Chd = os.getcwd()
file_path = 'BOM_List_OP.xlsx'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        dt_H1 = pd.read_csv(file_path)

except ValueError:
    ds1 = pd.read_excel(file_path, sheet_name="BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False) 
    dsn1 = pd.read_excel(file_path, sheet_name="Orginal_BOM", usecols=['PartNumber', 'Group', 'Priority', 'Long Des', 'RefList', 'Qty','Shape'],index_col=False)
    
    ds1 = ds1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
    ds1['RefList'] = ds1['RefList'].str.replace("_x000D_","")
    ds1['RefList'] = ds1['RefList'].str.replace(" ","")
    ds1['RefList'] = ds1['RefList'].str.replace("\n","")
    print(ds1)

#ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

    ds2 = ds1.explode('RefList')

    ds2['RefList'] = ds2['RefList'].str.replace(" "," ")

#ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

    print(ds2)

    ds2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

    ds2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

    ds2['B_Ref_List'] = ds2['B_Ref_List'] .str.strip('[]').str.split(',')

    print(ds2)

    ds2.to_dict()

    ds2.explode ('B_Ref_List',ignore_index=True)

    ds3 = ds2.explode('B_Ref_List',ignore_index=True) # split the Ref below example code

    '''import pandas as pd

    # Sample DataFrame
    data = {'ID': [1, 2], 'B_Ref_List': [['R1', 'R2'], ['R3', 'R4', 'R5']]}

    ds2 = pd.DataFrame(data)

    # Explode 'B_Ref_List'
    ds3 = ds2.explode('B_Ref_List', ignore_index=True)

    # Display the result
    print(ds3)
    Output:
        ID B_Ref_List
    0   1         R1
    1   1         R2
    2   2         R3
    3   2         R4
    4   2         R5'''

    ds2 = ds2[['Group','Priority','B_Part_No']]
    dc1 = ds2[['B_Part_No']]
    dc1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
    dc1['PBARPTN'] = dc1['PBARNO']
    dc1['PBARBAR'] = dc1['PBARNO']
    dc1.insert(3,'PBARQTY', 10000)
    dc1.insert(4,'PBARFTYP', 3)

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
    dfs2 = ds2[['Group','Priority','B_Part_No']]
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("15","A")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("14","B")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("13","C")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("12","D")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("11","E")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("10","F")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("9","PTN_9")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("8","PTN_8")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("7","PTN_7")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("6","PTN_6")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("5","PTN_5")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("4","PTN_4")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("3","PTN_3")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("2","PTN_2")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("1","PTN_1")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("0","PTN_0")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("A","PTN_15")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("B","PTN_14")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("C","PTN_13")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("D","PTN_12")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("E","PTN_11")
    dfs2['Priority'] = dfs2['Priority'].astype(str).str.replace("F","PTN_10")
    dfs2.dropna(subset=['Group'], inplace=True)
    #df2 = dfs2.pivot(index='Group',columns='Priority',values='B_Part_No')

        # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
    # Check for duplicate entries in 'Group' and 'Priority'
    duplicate_entries = dfs2[dfs2.duplicated(subset=['Group', 'Priority'], keep=False)]

    if not duplicate_entries.empty:
        # Show an error message if duplicates are found
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # If no duplicates, proceed with pivoting
    try:
        df2 = dfs2.pivot(index='Group', columns='Priority', values='B_Part_No')
    except ValueError as e:
        # Handle the exception gracefully
        error_message = f"An error occurred during pivoting: {e}"
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    ds3.head()

    T10_col = ds3.pop('B_Ref_List') # col-1

    ds3.insert(0, 'B_Ref_List', T10_col)

    ds3 = ds3[['B_Ref_List','B_Part_No','Long Des']]

    ds1.dropna(subset=['RefList'], inplace=True)
    ds3.dropna(subset=['B_Ref_List'], inplace=True)

#ONLY AVL PARTMASTER AND GOUPING
    dsn1 = dsn1[['PartNumber', 'Group','Priority','Long Des','Qty','Shape','RefList']]
    dsn1['RefList'] = dsn1['RefList'].str.replace("_x000D_","")
    dsn1['RefList'] = dsn1['RefList'].str.replace(" ","")
    dsn1['RefList'] = dsn1['RefList'].str.replace("\n","")
    print(dsn1)

#ds1 = pd.read_excel('Filename_OP.xlsx','BOM', index_col=False)

    dsn2 = dsn1.explode('RefList')

    dsn2['RefList'] = dsn2['RefList'].str.replace(" "," ")

#ds2.drop(ds2.iloc[:, 1:6], inplace=True, axis=1)

    print(dsn2)

    dsn2.rename(columns = {'PartNumber':'B_Part_No'}, inplace = True)

    dsn2.rename(columns = {'RefList':'B_Ref_List'}, inplace = True)

    dsn2['B_Ref_List'] = dsn2['B_Ref_List'] .str.strip('[]').str.split(',')

    dsn2.to_dict()

    dsn2.explode ('B_Ref_List',ignore_index=True)

    dsn3 = dsn2.explode('B_Ref_List',ignore_index=True)

    dsn2 = dsn2[['Group','Priority','B_Part_No']]

    # Condition: Check if Priority is only 0
    if (dsn2['Priority'] == 0).all():
    # Check if 1 and 2 are not present
        if not ((dsn2['Priority'] == 1) | (dsn2['Priority'] == 2)).any():
    # Add 1, 2, 3 in Priority column
            
            dsn2['Priority'] = dsn2['Priority']
            
    # Append corresponding Dummy_Part rows
            dummy_data = {'Group': ['B89P13', 'B89P13', 'B89P13'],
                        'Priority': [1, 2, 3],
                        'B_Part_No': ['Dummy_Part1', 'Dummy_Part2', 'Dummy_Part3']}
            
            dummy_df = pd.DataFrame(dummy_data)
            dsn2 = pd.concat([dsn2, dummy_df], ignore_index=True)

    # Continue with the rest of your code
    print(dsn2[['Group', 'Priority', 'B_Part_No']])

    dcn1 = dsn2[['B_Part_No']]
    duplicate_rows = dcn1[dcn1.duplicated(subset=['B_Part_No'], keep=False)]
    
    if not duplicate_rows.empty:
        # Show an error message if duplicates are found
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = f"Duplicate entries found in 'B_Part_No':\nCheck the BOM! PartNo Col.\n{duplicate_rows}"
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code
    
    print(dcn1)

    # Assuming 'B_Part_No' contains values like 'Dummy_Part1', 'Dummy_Part2', 'Dummy_Part3'
    dummy_values = [f'Dummy_Part{i}' for i in range(1, 4)]

    # Remove rows where 'B_Part_No' contains dummy values
    dcn1 = dcn1[~dcn1['B_Part_No'].isin(dummy_values)]

    dcn1.rename(columns = {'B_Part_No':'PBARNO'}, inplace = True)
    dcn1['PBARPTN'] = dcn1['PBARNO']
    dcn1['PBARBAR'] = dcn1['PBARNO']
    dcn1.insert(3,'PBARQTY', 10000)
    dcn1.insert(4,'PBARFTYP', 3)

    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>#
    dfsn2 = dsn2[['Group','Priority','B_Part_No']]
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("15","A")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("14","B")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("13","C")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("12","D")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("11","E")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("10","F")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("9","PTN_9")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("8","PTN_8")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("7","PTN_7")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("6","PTN_6")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("5","PTN_5")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("4","PTN_4")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("3","PTN_3")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("2","PTN_2")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("1","PTN_1")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("0","PTN_0")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("A","PTN_15")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("B","PTN_14")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("C","PTN_13")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("D","PTN_12")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("E","PTN_11")
    dfsn2['Priority'] = dfsn2['Priority'].astype(str).str.replace("F","PTN_10")
    dfsn2.dropna(subset=['Group'], inplace=True)
    #dfn2 = dfsn2.pivot(index='Group',columns='Priority',values='B_Part_No')

    # Assuming 'dfs2' is the DataFrame with 'Group', 'Priority', and 'B_Part_No' columns
    # Check for duplicate entries in 'Group' and 'Priority'
    duplicate_entries = dfsn2[dfsn2.duplicated(subset=['Group', 'Priority'], keep=False)]

    if not duplicate_entries.empty:
        # Show an error message if duplicates are found
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        error_message = f"Duplicate entries found in 'Group' and 'Priority':\n{duplicate_entries}"
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

    # If no duplicates, proceed with pivoting
    try:
        dfn2 = dfsn2.pivot(index='Group', columns='Priority', values='B_Part_No')
    except ValueError as e:
        # Handle the exception gracefully
        error_message = f"An error occurred during pivoting: {e}"
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showerror("Error", error_message)
        sys.exit(1)  # Exit the program with an error code

            # Desired column order
    desired_order = ['Group Name','PTN_1','PTN_2','PTN_3','PTN_4','PTN_5','PTN_6','PTN_7','PTN_8','PTN_9','PTN_10','PTN_11','PTN_12','PTN_13','PTN_14','PTN_15']
    #desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

    # Create a list of columns present in both DataFrame and desired_order
    common_columns = [col for col in desired_order if col in dfn2.columns]

    # Reorder the DataFrame based on the desired_order
    df_AL1 = dfn2[common_columns]

    '''    # Assuming df is your DataFrame
        column_to_check = 'PTN_15'

        # Check if the column is present
        if column_to_check not in df_AL1.columns:
            # Show a pop-up message if the column is not present
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            messagebox.showinfo("Notification", f"The column '{column_to_check}' is missing.")

        # Continue with the rest of your code
        print("Continuing with the rest of the code...")
        # Your next line of code here'''

    # Assuming df is your DataFrame
    column_to_check = 'PTN_11'

    # Check if the column is present
    if column_to_check in df_AL1.columns:
        # Show a pop-up message if the column is present
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        messagebox.showinfo("Notification", f"The column '{column_to_check}' is present.")

    # Continue with the rest of your code
    print("Continuing with the rest of the code...")
    # Your next line of code here

    dsn3.head()

    T10_col = dsn3.pop('B_Ref_List') # col-1

    dsn3.insert(0, 'B_Ref_List', T10_col)

    dsn3 = dsn3[['B_Ref_List','B_Part_No','Long Des']]

    dsn1.dropna(subset=['RefList'], inplace=True)
    dsn3.dropna(subset=['B_Ref_List'], inplace=True)

    #NEW PN# PARTNO
    ds1['PartNO'] = "PN#"
    ds1["PartNumber"] = ds1['PartNO'].astype(str) +""+ ds1['PartNumber'].astype(str)
    del ds1['PartNO']
    ds3['PartNO'] = "PN#"
    ds3["B_Part_No"] = ds3['PartNO'].astype(str) +""+ ds3['B_Part_No'].astype(str)
    del ds3['PartNO']

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx") as writer:

    #dt_H.to_excel(writer, sheet_name="Home", index=False)  
    #df1.to_excel(writer, sheet_name="FS_upload", index=False)
    #df2.to_excel(writer, sheet_name="FS_Count", index=False)
    #df3.to_excel(writer, sheet_name="FS-Program Split", index=False)
    #df3_1.to_excel(writer, sheet_name="FL_Upload", index=False)
    #df4.to_excel(writer, sheet_name="FL_Verify", index=False)
        ds1.to_excel(writer, sheet_name="BOM", index=False)
        dsn2.to_excel(writer, sheet_name="AVL GROUP", index=False)
        dcn1.to_excel(writer, sheet_name="Part Master", index=False)
        #dfn2.to_excel(writer, sheet_name="AVL_SHEET", index=True)
        df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=True)
        ds3.to_excel(writer, sheet_name="BOM_Data", index=False)
    #df2.to_excel(writer, sheet_name="AVL_SHEET", index=True) this line record upto 1 & 0
    #dc1.to_excel(writer, sheet_name="Part Master", index=False) this line record uoto 1 & 0 
    #ds2.to_excel(writer, sheet_name="AVL GROUP", index=False) this line record upto PTN1
    pass
    print('The file does not exist.')

#########################################################################################################################################################################
#########################################################################################################################################################################
    #@@ CRD Inspection @@#
    print('\n')
    print("\033[92;4m******CRD CHECK******\033[0m")
    print('\n')
#########################################################################################################################################################################
#########################################################################################################################################################################
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()
    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

    # Assuming 'BOM DATA' sheet contains a column named 'Bom Ref'
    df_bom_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM_Data")
    duplicates_bom_data = df_bom_data[df_bom_data.duplicated(subset='B_Ref_List', keep=False)]
    # Assuming 'XY DATA' sheet contains a column named 'R'
    #df_xy_data = pd.read_excel("BOM_List_OP.xlsx", sheet_name="XY DATA")

    # Function to check for duplicates and display an error message
    def check_and_show_duplicates(df, column_name, sheet_name):
        duplicates = df[df.duplicated(subset=column_name, keep=False)]
        if not duplicates.empty:
            print("Duplicate values in 'B_Ref_List' column of 'BOM_Data':")
            print(duplicates_bom_data['B_Ref_List'].tolist())
            root = tk.Tk()
            root.withdraw()
            message = f"Duplicate values found in '{column_name}' column of '{sheet_name}' sheet:\n{duplicates[column_name].tolist()}"
            messagebox.showerror("Error", message)
            sys.exit()

    # Check for duplicates in 'Bom Ref' column of 'BOM DATA'
    check_and_show_duplicates(df_bom_data, 'B_Ref_List', 'BOM_Data')

    # Check for duplicates in 'R' column of 'XY DATA'
    #check_and_show_duplicates(df_xy_data, 'R', 'XY DATA')

#########################################################################################################################################################################
#########################################################################################################################################################################

##########################################################################################################################################
##########################################################################################################################################
    #@@ AVL Inspection @@#
##########################################################################################################################################

print('\n')
print("\033[92;4m******AVL LINE INSPECTION******\033[0m")
print('\n')

##########################################################################################################################################

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
file_path = 'BOM_List_OP.xlsx'
directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified'

print(os.path.isfile(file_path))
print(os.path.isfile(directory_path))

try:
    if os.path.isfile(file_path):
        df_Iav1 = pd.read_csv(file_path)

except ValueError:
    df_Iav1 = pd.read_excel(file_path, sheet_name="AVL_SHEET", index_col=False) 

    print(df_Iav1)

    # Function to check for missing values between two strings
def check_missing_values(row):
    start_index = None
    end_index = None

    for i in range(1, len(row) + 1):  # Check up to the last column
        col_name = f'PTN_{i}'
        if col_name in row.index:  # Check if the column exists
            current_value = row[col_name]

            if pd.isna(current_value):
                if start_index is None:
                    start_index = i
                end_index = i
            else:
                if start_index is not None and end_index is not None:
                    show_error(row['Group'], start_index, end_index)
                    start_index = None
                    end_index = None

# Function to show pop-up error message
def show_error(group, start_index, end_index):
    root = tk.Tk()
    root.withdraw()
    error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'."
    messagebox.showerror("Error", error_message)

# Check for missing values row-wise
for index, row in df_Iav1.iterrows():
    check_missing_values(row)

# Display the DataFrame with styling
print(df_Iav1)

# Function to check for missing values between two strings
def check_missing_values(row):
    start_index = None
    end_index = None

    for i in range(1, len(row) + 1):  # Check up to the last column
        col_name = f'PTN_{i}'
        if col_name in row.index:  # Check if the column exists
            current_value = row[col_name]

            if pd.isna(current_value):
                if start_index is None:
                    start_index = i
                end_index = i
            else:
                if start_index is not None and end_index is not None:
                    show_error(row.get('Group', 'Unknown Group'), start_index, end_index)
                    start_index = None
                    end_index = None

# Function to show pop-up error message
def show_error(group, start_index, end_index):
    root = tk.Tk()
    root.withdraw()
    
    error_message = f"Error: Missing values between PTN_{start_index} and PTN_{end_index} in group '{group}'.\nDo you want to stop the program?"
    response = messagebox.askquestion("Error", error_message)

    if response == 'yes':
        sys.exit(1)

# ...

# Check for missing values row-wise
for index, row in df_Iav1.iterrows():
    check_missing_values(row)

# Display the DataFrame with styling
print(df_Iav1)

##########################################################################################################################################

##########################################################################################################################################

#bil6 = pyfiglet.figlet_format("Part Master Process", width = 100)
print('\n')
print("\033[92;4m******Part Master Process******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################
    
#PART MASTER

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

# Excel file path
df_PM1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Part Master")
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx") as writer:
    df_PM1.to_excel(writer, sheet_name="T_PBAR", index=False)
excel_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx'

# Access database connection parameters
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_Part Master/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb')
access_db_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb'
driver = 'Microsoft Access Driver (*.mdb, *.accdb)'
user = ''
password = ''

# Set up the connection string
conn_str = f"DRIVER={{{driver}}};DBQ={access_db_path};UID={user};PWD={password};"

# Connect to the Access database
print(pyodbc.drivers())
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()

# Read Excel data into a pandas DataFrame
print('Open Excel....')
df = pd.read_excel(excel_file_path)
print(df.head(10))

# Define the table name in the Access database
print('open MS Access....')
table_name = 'T_PBAR'

# Check if the table exists
existing_tables = [table[2] for table in cursor.tables(tableType='TABLE')]
if table_name in existing_tables:
    # Append data to the existing table
    for _, row in df.iterrows():
        insert_query = f'''
        INSERT INTO {table_name} ({', '.join(df.columns)})
        VALUES ({', '.join(map(lambda x: f"'{row[x]}'", df.columns))})
        '''
        cursor.execute(insert_query)
        conn.commit()
        print('writing to access')
else:
    print(f"The table '{table_name}' does notYT exist in the Access database.")

# Close the database connection
conn.close()
print('write complete')

##########################################################################################################################################

##########################################################################################################################################

#bil7 = pyfiglet.figlet_format("AVL Progress", width = 100)
print('\n')
print("\033[92;4m******AVL Progress******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################

#AVL#@@#

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

print(pyodbc.drivers())

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='AVL_SHEET')

print(df1.head(10))

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df1.to_excel(writer, sheet_name="AVL_SHEET", index=False)

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
df_AL1 = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')

#new_column_name  = df_AL1.insert(1,'Group Name', '') #new_column_name = ('Group Name') # Get user input for the new column name
# Desired column name
desired_column_name = 'Group'

# Get user input for the new column name
new_column_name = ("Group Name")

# Get user input for the new column value
new_column_value = input(f"\033[93mEnter the value for the new column '{new_column_name}': \033[0m")
#dL1 = new_column_value
# Check if the desired column name exists
if desired_column_name in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value) #dL1 = new_column_value

print(df_AL1)

try:

    df_AL1['AVL Name']=df_AL1['PTN_1']
    # Replace values in 'AVL Name' with values from 'PTN_1' where 'PTN_1' is not empty
    #df_AL1['AVL Name'] = df_AL1['PTN_1'].fillna(df_AL1['AVL Name'])

except Exception as e:
    # Handle the exception gracefully
    error_message = f"An error occurred:\nSomething went wrong while assigning AVL values {e}"
    error_msg1 = f"Check SF-02 is deleted\nCheck AVL Priority assign Properly {e}"
    # Show error message in a pop-up box
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    messagebox.showerror("Error", error_message)
    messagebox.showerror("AVL Error", error_msg1)
    #sys.exit(1)  # Exit the program with an error code

first_column = df_AL1.pop('AVL Name')

df_AL1.insert(2, 'AVL Name', first_column)

df_AL1.insert(3, 'Comment', '')

df_AL1['Comment'] = df_AL1['Group']

#PTN_1>> numpy changes #np.NaN > np.nan 2.0

# Desired column name
desired_column_name1 = 'PTN_1'

# New column to insert
new_column_name = 'P_1'
new_column_value = '1'

# Check if the desired column name exists
if desired_column_name1 in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name1)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

#PTN_2>>

desired_column_name2 = 'PTN_2'

# New column to insert
new_column_name = 'P_2'
new_column_value = '0'

# Check if the desired column name exists
if desired_column_name2 in df_AL1.columns:
    # Find the index of the desired column
    index_of_desired_column = df_AL1.columns.get_loc(desired_column_name2)
    
    # Insert the new column next to the desired column
    df_AL1.insert(index_of_desired_column + 1, new_column_name, new_column_value)

#PTN_3>>

desired_column_name3 = 'PTN_3'
column_to_replace3 = 'PTN_3'

# Check if the column exists in the DataFrame
if column_to_replace3 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name3 = 'P_3'
    new_column_value3 = '0'

    # Check if the desired column name exists
    if desired_column_name3 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name3)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name3,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name3].any():
            print(f"Values are present in '{desired_column_name3}' column:")
            print(df_AL1[desired_column_name3])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name3] != "$", new_column_name3] = 0
        else:
            print(f"No values are present in '{desired_column_name3}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name3}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

column_to_replace3 = 'PTN_3'

# Check if the column exists in the DataFrame
if column_to_replace3 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace3] = df_AL1[column_to_replace3].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace3}' does not exist in the DataFrame.")

#PTN_4>>

desired_column_name4 = 'PTN_4'
column_to_replace4 = 'PTN_4'

# Check if the column exists in the DataFrame
if column_to_replace4 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name4 = 'P_4'
    new_column_value4 = '0'

    # Check if the desired column name exists
    if desired_column_name4 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name4)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name4,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name4].any():
            print(f"Values are present in '{desired_column_name4}' column:")
            print(df_AL1[desired_column_name4])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name4] != "$", new_column_name4] = 0
        else:
            print(f"No values are present in '{desired_column_name4}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name4}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

column_to_replace4 = 'PTN_4'

# Check if the column exists in the DataFrame
if column_to_replace4 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace4] = df_AL1[column_to_replace4].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace4}' does not exist in the DataFrame.")

#PTN_5>>

desired_column_name5 = 'PTN_5'
column_to_replace5 = 'PTN_5'

# Check if the column exists in the DataFrame
if column_to_replace5 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name5 = 'P_5'
    new_column_value5 = '0'

    # Check if the desired column name exists
    if desired_column_name5 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name5)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name5,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name5].any():
            print(f"Values are present in '{desired_column_name5}' column:")
            print(df_AL1[desired_column_name5])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name5] != "$", new_column_name5] = 0
        else:
            print(f"No values are present in '{desired_column_name5}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name5}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

column_to_replace5 = 'PTN_5'

# Check if the column exists in the DataFrame
if column_to_replace5 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace5] = df_AL1[column_to_replace5].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace5}' does not exist in the DataFrame.")

#PTN_6>>

desired_column_name6 = 'PTN_6'
column_to_replace6 = 'PTN_6'

# Check if the column exists in the DataFrame
if column_to_replace6 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name6 = 'P_6'
    new_column_value6 = '0'

    # Check if the desired column name exists
    if desired_column_name6 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name6)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name6,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name6].any():
            print(f"Values are present in '{desired_column_name6}' column:")
            print(df_AL1[desired_column_name6])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name6] != "$", new_column_name6] = 0
        else:
            print(f"No values are present in '{desired_column_name6}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name6}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

column_to_replace6 = 'PTN_6'

# Check if the column exists in the DataFrame
if column_to_replace6 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace6] = df_AL1[column_to_replace6].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace6}' does not exist in the DataFrame.")

#PTN_7>>

desired_column_name7 = 'PTN_7'
column_to_replace7 = 'PTN_7'

# Check if the column exists in the DataFrame
if column_to_replace7 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name7 = 'P_7'
    new_column_value7 = '0'

    # Check if the desired column name exists
    if desired_column_name7 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name7)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name7,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name7].any():
            print(f"Values are present in '{desired_column_name7}' column:")
            print(df_AL1[desired_column_name7])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name7] != "$", new_column_name7] = 0
        else:
            print(f"No values are present in '{desired_column_name7}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name7}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

column_to_replace7 = 'PTN_7'

# Check if the column exists in the DataFrame
if column_to_replace7 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace7] = df_AL1[column_to_replace7].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace7}' does not exist in the DataFrame.")

#PTN_8>>

desired_column_name8 = 'PTN_8'
column_to_replace8 = 'PTN_8'

# Check if the column exists in the DataFrame
if column_to_replace8 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name8 = 'P_8'
    new_column_value8 = '0'

    # Check if the desired column name exists
    if desired_column_name8 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name8)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name8,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name8].any():
            print(f"Values are present in '{desired_column_name8}' column:")
            print(df_AL1[desired_column_name8])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name8] != "$", new_column_name8] = 0
        else:
            print(f"No values are present in '{desired_column_name8}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name8}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

column_to_replace8 = 'PTN_8'

# Check if the column exists in the DataFrame
if column_to_replace8 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace8] = df_AL1[column_to_replace8].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace8}' does not exist in the DataFrame.")

#PTN_9>>

desired_column_name9 = 'PTN_9'
column_to_replace9 = 'PTN_9'

# Check if the column exists in the DataFrame
if column_to_replace9 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name9 = 'P_9'
    new_column_value9 = '0'

    # Check if the desired column name exists
    if desired_column_name9 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name9)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name9,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name9].any():
            print(f"Values are present in '{desired_column_name9}' column:")
            print(df_AL1[desired_column_name9])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name9] != "$", new_column_name9] = 0
        else:
            print(f"No values are present in '{desired_column_name9}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name9}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

column_to_replace9 = 'PTN_9'

# Check if the column exists in the DataFrame
if column_to_replace9 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace9] = df_AL1[column_to_replace9].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace9}' does not exist in the DataFrame.")

#PTN_10>>

desired_column_name10 = 'PTN_10'
column_to_replace10 = 'PTN_10'

# Check if the column exists in the DataFrame
if column_to_replace10 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name10 = 'P_10'
    new_column_value10 = '0'

    # Check if the desired column name exists
    if desired_column_name10 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name10)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name10,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name10].any():
            print(f"Values are present in '{desired_column_name10}' column:")
            print(df_AL1[desired_column_name10])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name10] != "$", new_column_name10] = 0
        else:
            print(f"No values are present in '{desired_column_name10}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name10}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

column_to_replace10 = 'PTN_10'

# Check if the column exists in the DataFrame
if column_to_replace10 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace10] = df_AL1[column_to_replace10].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace10}' does not exist in the DataFrame.")

#PTN_11>>

desired_column_name11 = 'PTN_11'
column_to_replace11 = 'PTN_11'

# Check if the column exists in the DataFrame
if column_to_replace11 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name11 = 'P_11'
    new_column_value11 = '0'

    # Check if the desired column name exists
    if desired_column_name11 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name11)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name11,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name11].any():
            print(f"Values are present in '{desired_column_name11}' column:")
            print(df_AL1[desired_column_name11])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name11] != "$", new_column_name11] = 0
        else:
            print(f"No values are present in '{desired_column_name11}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name11}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

column_to_replace11 = 'PTN_11'

# Check if the column exists in the DataFrame
if column_to_replace11 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace11] = df_AL1[column_to_replace11].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace11}' does not exist in the DataFrame.")

#PTN_12>>

desired_column_name12 = 'PTN_12'
column_to_replace12 = 'PTN_12'

# Check if the column exists in the DataFrame
if column_to_replace12 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name12 = 'P_12'
    new_column_value12 = '0'

    # Check if the desired column name exists
    if desired_column_name12 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name12)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name12,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name12].any():
            print(f"Values are present in '{desired_column_name12}' column:")
            print(df_AL1[desired_column_name12])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name12] != "$", new_column_name12] = 0
        else:
            print(f"No values are present in '{desired_column_name12}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name12}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

column_to_replace12 = 'PTN_12'

# Check if the column exists in the DataFrame
if column_to_replace12 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace12] = df_AL1[column_to_replace12].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace12}' does not exist in the DataFrame.")

#PTN_13>>

desired_column_name13 = 'PTN_13'
column_to_replace13 = 'PTN_13'

# Check if the column exists in the DataFrame
if column_to_replace13 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name13 = 'P_13'
    new_column_value13 = '0'

    # Check if the desired column name exists
    if desired_column_name13 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name13)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name13,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name13].any():
            print(f"Values are present in '{desired_column_name13}' column:")
            print(df_AL1[desired_column_name13])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name13] != "$", new_column_name13] = 0
        else:
            print(f"No values are present in '{desired_column_name13}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name13}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

column_to_replace13 = 'PTN_13'

# Check if the column exists in the DataFrame
if column_to_replace13 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace13] = df_AL1[column_to_replace13].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace13}' does not exist in the DataFrame.")

#PTN_14>>

desired_column_name14 = 'PTN_14'
column_to_replace14 = 'PTN_14'

# Check if the column exists in the DataFrame
if column_to_replace14 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name14 = 'P_14'
    new_column_value14 = '0'

    # Check if the desired column name exists
    if desired_column_name14 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name14)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name14,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name14].any():
            print(f"Values are present in '{desired_column_name14}' column:")
            print(df_AL1[desired_column_name14])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name14] != "$", new_column_name14] = 0
        else:
            print(f"No values are present in '{desired_column_name14}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name14}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

column_to_replace14 = 'PTN_14'

# Check if the column exists in the DataFrame
if column_to_replace14 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace14] = df_AL1[column_to_replace14].replace("$", np.nan)
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace14}' does not exist in the DataFrame.")

#PTN_15>>

desired_column_name15 = 'PTN_15'
column_to_replace15 = 'PTN_15'

# Check if the column exists in the DataFrame
if column_to_replace15 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace(np.nan, "$")
    print("\nDataFrame after replacement:")
    print(df_AL1)

    # New column to insert
    new_column_name15 = 'P_15'
    new_column_value15 = '0'

    # Check if the desired column name exists
    if desired_column_name15 in df_AL1.columns:
        # Find the index of the desired column
        index_of_desired_column = df_AL1.columns.get_loc(desired_column_name15)

        # Insert the new column next to the desired column
        df_AL1.insert(index_of_desired_column + 1, new_column_name15,  '')

        # Check if values are present in the desired column
        if df_AL1[desired_column_name15].any():
            print(f"Values are present in '{desired_column_name15}' column:")
            print(df_AL1[desired_column_name15])

        # Set the values in the new column to zero for cells where values are present in the desired column
            df_AL1.loc[df_AL1[desired_column_name15] != "$", new_column_name15] = 0
        else:
            print(f"No values are present in '{desired_column_name15}' column.")

        print("\nUpdated DataFrame:")
        print(df_AL1)
    else:
        print(f"Column '{desired_column_name15}' does not exist in the DataFrame.")
else:
    print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

column_to_replace15 = 'PTN_15'

# Check if the column exists in the DataFrame
if column_to_replace15 in df_AL1.columns:
    # Replace NaN values with "$" in the specified column
    df_AL1[column_to_replace15] = df_AL1[column_to_replace15].replace("$", np.nan) #np.NaN > np.nan
    print("\nDataFrame after replacement:")
    print(df_AL1)
else:
    print(f"Column '{column_to_replace15}' does not exist in the DataFrame.")

print(df_AL1.head(5))

del df_AL1['Group']

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

# Example DataFrame
data = pd.read_excel('AVL.xlsx', sheet_name='AVL_SHEET')
df_AL1 = pd.DataFrame(data)

# Desired column order
desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10']
#desired_order = ['Group Name','AVL Name','Comment','PTN_1','P_1','PTN_2','P_2','PTN_3','P_3','PTN_4','P_4','PTN_5','P_5','PTN_6','P_6','PTN_7','P_7','PTN_8','P_8','PTN_9','P_9','PTN_10','P_10','PTN_11','P_11','PTN_12','P_12','PTN_13','P_13','PTN_14','P_14','PTN_15','P_15']

# Create a list of columns present in both DataFrame and desired_order
common_columns = [col for col in desired_order if col in df_AL1.columns]

# Reorder the DataFrame based on the desired_order
df_AL1 = df_AL1[common_columns]

# Display the reordered DataFrame
print(df_AL1)

with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.xlsx") as writer:
    df_AL1.to_excel(writer, sheet_name="AVL_SHEET", index=False)
    df_AL1.T.reset_index().T.to_excel(writer, sheet_name="AVL_SHEET", header=False ,index=False)
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')

#read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx',skiprows=0)

read_file = pd.read_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.xlsx')

read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', index = None, header= None)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()

# Replace '0.0' with '0' in the content
modified_content = content.replace('0.0', '0')

# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(modified_content)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC20 = content.replace(',,,,,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC20)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC19 = content.replace(',,,,,,,,,,,,,,,,,,,', '')
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC19)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC18 = content.replace(',,,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC18)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC17 = content.replace(',,,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC17)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC16 = content.replace(',,,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC16)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC15 = content.replace(',,,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC15)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC14 = content.replace(',,,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC14)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC13 = content.replace(',,,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC13)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC12 = content.replace(',,,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC12)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC11 = content.replace(',,,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC11)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC10 = content.replace(',,,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC10)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC9 = content.replace(',,,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC9)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC8 = content.replace(',,,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC8)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC7 = content.replace(',,,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC7)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC6 = content.replace(',,,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC6)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC5 = content.replace(',,,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC5)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC4 = content.replace(',,,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC4)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC3 = content.replace(',,,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC3)

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC2 = content.replace(',,', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC2)
    #output_file.write('MC2','MC3','MC4','MC5','MC6','MC7','MC8','MC9','MC10','MC11','MC12','MC13','MC14','MC15','MC16','MC17','MC18','MC19','MC20')

'''# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC1 = content.replace('+', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC1)'''

# Specify the path to your text file
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'

# Read the content of the text file
with open(txt_file_path, 'r') as file:
    content = file.read()
MC0 = content.replace('.0', '')
# Write the modified content back to the text file
with open('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt', 'w') as output_file:
    output_file.write(MC0)

print(f"AVL CREATED: D:/NX_BACKWORK/r'AVL.txt")

# Specify the path to your text file
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
txt_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.txt'
df = pd.read_table(txt_file_path, delimiter='\t', quoting=3)  # 3 corresponds to QUOTE_NONE
csv_file_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv'
df.to_csv(csv_file_path, index=False, sep='\t')  # 0 corresponds to QUOTE_NONE

#read_file = pd.read_table (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.txt', sep='"')
#read_file.to_excel (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVList.xlsx', index=None)
#read_file.to_csv (r'D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Verified\AVL.csv', index = None, header= None)


##########################################################################################################################################

##########################################################################################################################################

#bil8 = pyfiglet.figlet_format("Feeder Verification", width = 100)
print('\n')
print("\033[92;4m******Feeder Verification******\033[0m")
print('\n')
##########################################################################################################################################

##########################################################################################################################################

#FEEDER VERIFICATION CODE BOM AND FEEDER AS VISE VERSA

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df1 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='BOM_Data')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
df2 = pd.read_excel('FeederSetup.xlsx', sheet_name='FeederCol')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
df111 = pd.read_excel('BOM_List_OP.xlsx', sheet_name='BOM')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
xls = pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
df112 = pd.read_excel('FeederSetup.xlsx', sheet_name='FeederSetup')
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:
    df1.to_excel(writer, sheet_name="BOM_Data", index=False)
    df2.to_excel(writer, sheet_name="FeederCol", index=False)
    df111.to_excel(writer, sheet_name="BOM", index=False)
    df112.to_excel(writer, sheet_name="FeederSetup", index=False)

xls = pd.ExcelFile('FeederVerify.xlsx',engine='openpyxl')
df1 = pd.read_excel("FeederVerify.xlsx", sheet_name='BOM_Data')
df2 = pd.read_excel("FeederVerify.xlsx", sheet_name='FeederCol')
df111 = pd.read_excel("FeederVerify.xlsx", sheet_name='BOM')
df112 = pd.read_excel("FeederVerify.xlsx", sheet_name='FeederSetup')
    
df2['Feeder Reference'] = df2['F_Ref_List']
df1.rename(columns = {'B_Ref_List':'F_Ref_List'}, inplace = True)
    #df1['B_Ref.List'] = df1['F_Ref_List']
df3 = pd.merge(df1 , df2, on='F_Ref_List', how='left')
df3.rename(columns = {'F_Ref_List':'BOM Reference'}, inplace = True)
df1.rename(columns = {'F_Ref_List':'B_Ref_List'}, inplace = True)
print(df1,df2)

df111.rename(columns = {'PartNumber':'F_Part_No'}, inplace = True)
df111 = df111[['F_Part_No','Long Des']]
df113 = pd.merge(df111 , df112, on='F_Part_No', how='inner') # Merge on 'F_Part_No'
df113.rename(columns = {'F_Part_No':'Part Number'}, inplace = True)
df113.rename(columns = {'Location':'Feeder Location'}, inplace = True)
df113.rename(columns = {'Long Des':'Part Description'}, inplace = True)
df113.rename(columns = {'F_Ref_List':'Reference'}, inplace = True)
df113 = df113[['Feeder Location','FeederName','Type','Size','FeedPitch','Part Height','Part Number','Part Description','Reference','QTY','Side','ModelName']]

df3["BOM and Feeder Compare"] = (df3["B_Part_No"] == df3["F_Part_No"])
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].replace('TRUE','MATCH')
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].replace('FALSE','MISS_MATCH')

df3 = df3.copy()
df3['BOM and Feeder Compare'] = df3['BOM and Feeder Compare'].map({True: 'Match', False: 'Miss_Match'})
df3.sort_values(by='BOM and Feeder Compare', inplace=True, ascending=False)

df4 = df3['BOM and Feeder Compare'].value_counts()
#df4 = df3['Size'].value_counts()
df5 = df3['Side'].value_counts()
#df6 = df3['F_Ref_List'].value_counts()
#df7 = df3['B_Part_No'].value_counts() #LINE MOVED BELOW
#df8 = df3['B_Ref.List'].value_counts() 
#df9 = df3['F_Part_No'].value_counts() #LINE MOVED BELOW
#df10 = df3['FeederName'].value_counts()
df11 = len(df1['B_Ref_List'])
print(f'Total count of rows in the "B_Ref_List" column: {df11}')
df12 = len(df2['F_Ref_List'])
print(f'Total count of rows in the "F_Ref_List" column: {df12}')
print('***')
dbf1 = print(df1)
print('***')
dbf2 = print(df2)
print('***')
dbf2 = df2.copy()
dbf1 = df1.copy()
dbf2_col = dbf2.pop('Feeder Reference')
dbf2.insert(1, 'Feeder Reference', dbf2_col)
dbf2.rename(columns = {'F_Ref_List':'B_Ref_List'}, inplace = True)
dbf3 = pd.merge(dbf2 , dbf1, on='B_Ref_List', how='left')

dbf3["Feeder and BOM Compare"] = (dbf3["F_Part_No"] == dbf3["B_Part_No"])
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].replace('TRUE','MATCH')
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].replace('FALSE','MISS_MATCH')

dbf3 = dbf3.copy()
dbf3['Feeder and BOM Compare'] = dbf3['Feeder and BOM Compare'].map({True: 'Match', False: 'Miss_Match'})
dbf3.sort_values(by='Feeder and BOM Compare', inplace=True, ascending=False)

dbf4 = dbf3['Feeder and BOM Compare'].value_counts()

#NEW PN# PART NO

df3['F_Part_No'] = df3['F_Part_No'].str.replace('PN#','')
df3['B_Part_No'] = df3['B_Part_No'].str.replace('PN#','')

dbf3['F_Part_No'] = dbf3['F_Part_No'].str.replace('PN#','')
dbf3['B_Part_No'] = dbf3['B_Part_No'].str.replace('PN#','')


# Define a function for row styling
def highlight_row(row):
    return ['background-color: lightgreen' if 'Match' in row.values else
            'background-color: yellow' if 'Miss_Match' in row.values else
            '' for _ in row]

# Apply the styling function to the DataFrame
styled_df3 = df3.style.apply(highlight_row, axis=1)

# Define a function for row styling
def highlight_row(row):
    return ['background-color: lightgreen' if 'Match' in row.values else
            'background-color: yellow' if 'Miss_Match' in row.values else
            '' for _ in row]

# Apply the styling function to the DataFrame
styled_dbf3 = dbf3.style.apply(highlight_row, axis=1)

#NEW PN# PART NO
df113['Part Number'] = df113['Part Number'].str.replace('PN#','')
df1['B_Part_No'] = df1['B_Part_No'].str.replace('PN#','')
df2['F_Part_No'] = df2['F_Part_No'].str.replace('PN#','')

df7 = df3['B_Part_No'].value_counts()
df9 = df3['F_Part_No'].value_counts()

# Save the styled DataFrame to Excel
with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx") as writer:

        styled_df3.to_excel(writer, sheet_name="Verify_data_BAF", index=False, engine='openpyxl')
        df4.to_excel(writer, sheet_name="BOM and Feeder Compare",index=TRUE)
        #dbf2.to_excel(writer, sheet_name="Verify_data_FAB", index=False)
        styled_dbf3.to_excel(writer, sheet_name="Verify_data_FAB", index=False, engine='openpyxl')
        dbf4.to_excel(writer, sheet_name="Feeder and BOM Compare",index=TRUE)
        df113.to_excel(writer, sheet_name="Upload_data", index=False)
        df1.to_excel(writer, sheet_name="BOM_data", index=False)
        df2.to_excel(writer, sheet_name="Feeder_data", index=False)
        df5.to_excel(writer, sheet_name="Side",index=TRUE)
        #df6.to_excel(writer, sheet_name="F_Ref_List",index=TRUE)
        df7.to_excel(writer, sheet_name="B_Part_No",index=TRUE)
        #df8.to_excel(writer, sheet_name="B_Ref.List",index=TRUE)
        df9.to_excel(writer, sheet_name="F_Part_No",index=TRUE)
        #df10.to_excel(writer, sheet_name="FeederName",index=TRUE)
        # Save the count to an Excel file
        count_df = pd.DataFrame({'BOM_Data Ref, Count': [df11]})
        count_df.to_excel(writer, sheet_name="BOM Count", index=TRUE)
        count_df = pd.DataFrame({'Feeder_Data Ref, Count': [df12]})
        count_df.to_excel(writer, sheet_name="Feeder Count", index=TRUE)
##########################################################################################################################################

##########################################################################################################################################
    #CHANGE THE PN# IN EVERY PART NO#  BOM_List_OP & 
os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls=pd.ExcelFile('BOM_List_OP.xlsx',engine='openpyxl')
dfblop1 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM")
dfblop2 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="AVL GROUP")
dfblop3 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="Part Master")
dfblop4 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="AVL_SHEET")
dfblop5 = pd.read_excel("BOM_List_OP.xlsx", sheet_name="BOM_Data")

dfblop1['PartNumber'] = dfblop1['PartNumber'].str.replace('PN#','')
dfblop5['B_Part_No'] = dfblop5['B_Part_No'].str.replace('PN#','')

with pd.ExcelWriter('BOM_List_OP.xlsx', engine='openpyxl', mode='w') as writer:
    dfblop1.to_excel(writer, sheet_name="BOM", index=False)
    dfblop2.to_excel(writer, sheet_name="AVL GROUP", index=False)
    dfblop3.to_excel(writer, sheet_name="Part Master", index=False)
    dfblop4.to_excel(writer, sheet_name="AVL_SHEET", index=False)
    dfblop5.to_excel(writer, sheet_name="BOM_Data", index=False)

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls=pd.ExcelFile('FeederSetup.xlsx',engine='openpyxl')
dffstb11 = pd.read_excel("FeederSetup.xlsx", sheet_name="Home")
dffstb12 = pd.read_excel("FeederSetup.xlsx", sheet_name="FeederSetup")
dffstb13 = pd.read_excel("FeederSetup.xlsx", sheet_name="FeederCol")
dffstb14 = pd.read_excel("FeederSetup.xlsx", sheet_name="FeederSize")
dffstb15 = pd.read_excel("FeederSetup.xlsx", sheet_name="Total side Count")
dffstb16 = pd.read_excel("FeederSetup.xlsx", sheet_name="FeederName")
dffstb17 = pd.read_excel("FeederSetup.xlsx", sheet_name="Type")

dffstb12['F_Part_No'] = dffstb12['F_Part_No'].str.replace('PN#','')
dffstb13['F_Part_No'] = dffstb13['F_Part_No'].str.replace('PN#','')

with pd.ExcelWriter('FeederSetup.xlsx', engine='openpyxl', mode='w') as writer:
    dffstb11.to_excel(writer, sheet_name="Home", index=False)
    dffstb12.to_excel(writer, sheet_name="FeederSetup", index=False)
    dffstb13.to_excel(writer, sheet_name="FeederCol", index=False)
    dffstb14.to_excel(writer, sheet_name="FeederSize", index=False)
    dffstb15.to_excel(writer, sheet_name="Total side Count", index=False)
    dffstb16.to_excel(writer, sheet_name="FeederName", index=False)
    dffstb17.to_excel(writer, sheet_name="Type", index=False)

    
##########################################################################################################################################

##########################################################################################################################################

#bil9 = pyfiglet.figlet_format("FeederSetup Verification Result", width = 200)
print('\n')
print("\033[92;4m******FeederSetup Verification Result******\033[0m")
print('\n')
for i in range(100):
    row = "="*i + ">"
    sys.stdout.write("%s\r %d%%\r" %(row, i + 1))
    sys.stdout.flush()
    time.sleep(0.1)

##########################################################################################################################################

##########################################################################################################################################

print("\033[1;92;4mFeederSetup_Verification__Compelete $ PROCESS $\033[0m")

print('\n')

current_datetime = datetime.now()

# Print the current date
print("\033[35mCurrent Date:\033[0m", current_datetime.date())

# Print the current time
print("\033[35mCurrent Time:\033[0m", current_datetime.time())

os.getcwd()
Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

xls=pd.ExcelFile('FeederVerify.xlsx',engine='openpyxl')
df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM_data")
dfs1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder_data")
df2 = pd.read_excel("FeederVerify.xlsx", sheet_name="Side")
df3 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM and Feeder Compare")
df4 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder and BOM Compare")
dfs21 = pd.read_excel('Feederverify.xlsx', sheet_name="Feeder_data", usecols=['Location','F_Part_No','FeederName','Type','Size','FeedPitch','Part Height','Status','QTY','Side','ModelName','F_Ref_List','Feeder Reference'],index_col=False)
dfs22 = pd.read_excel('Feederverify.xlsx', sheet_name="BOM_data", usecols=['B_Ref_List','B_Part_No','Long Des'],index_col=False)
dfs3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_BAF")
dbf3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_FAB")
dfsg21 = dfs21[dfs21['Feeder Reference'].duplicated() == True]
dfsg22 = dfs22[dfs22['B_Ref_List'].duplicated() == True]
dfsg31 = dfs3[dfs3['BOM and Feeder Compare'].str.contains('Miss_Match')]
dfsg32 = dbf3[dbf3['Feeder and BOM Compare'].str.contains('Miss_Match')]

print('\n')

# Print the formatted date and time
print(f"\033[35mDate and Time: {formatted_datetime}\033[0m")

print('\n')

rc = len(df1)
rc1 = len(dfs1)

print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print(Chd,'\\__BOM__\\',dL1)
print(Chd,'\\__FeederSetup__\\',dL2)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
ds1 = print("BOM Count:",rc)
print('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx, Sheetname=BOM_data')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
ds9 = print("Feeder Count:",rc1)
print('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx, Sheetname=Feedercol')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
ds2 = print("BOT & TOP Count:")
ds2 = print(df2)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
ds3 = print("Compare Count:")
ds3 = print(df3)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
ds4 = print("Compare Count:")
ds4 = print(df4)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print("Feeder duplicate Reference")
print(dfsg21)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print("BOM duplicate Reference")
print(dfsg22)
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print("Miss Match Row BOM to Feeder")
print(dfsg31) 
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print('\n')
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
print("Miss Match Row Feeder to BOM")
print(dfsg32) 
print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

import tkinter as tk
from tkinter import scrolledtext, messagebox
import pandas as pd
from datetime import datetime
import os
import sys
import time

class FeederSetupApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FeederSetup")

        # Log window
        self.log_window = scrolledtext.ScrolledText(
            self.root, width=160, height=40, font=('Courier', 9), wrap=tk.WORD
        )
        self.log_window.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

        # Buttons
        self.save_button = tk.Button(self.root, text="Save to Excel", command=self.save_to_excel)
        self.save_button.grid(row=1, column=0, sticky='e', padx=10, pady=(0, 10))

        self.quit_button = tk.Button(self.root, text="Quit", command=self.root.quit)
        self.quit_button.grid(row=1, column=1, sticky='w', padx=10, pady=(0, 10))

        # Initial log print
        self.print_to_log("FeederSetup_Verification_Result_Compelete $ PROCESS $")
        sys.stdout.write("\n")

        # Run processing and display data
        self.process_data()

    def print_to_log(self, *args):
        message = ' '.join(map(str, args))
        self.log_window.insert(tk.END, message + '\n')
        self.log_window.see(tk.END)

    def save_to_excel(self):
        log_contents = self.log_window.get("1.0", tk.END).strip()
        if not log_contents:
            messagebox.showinfo("Empty Log", "No content to save.")
            return

        current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S") # %Y-%m-%d %I:%M:%S %p
        lines = log_contents.split('\n')
        df = pd.DataFrame({'LogContents': lines})

        excel_file_path = 'FeederVerify.xlsx'
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
            df.to_excel(writer, sheet_name=f'Log_{current_datetime}', index=False)

        messagebox.showinfo("Success", f"Log saved to {excel_file_path}, Sheet: Log_{current_datetime}")

    def process_data(self):
        formatted_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
        Chd = os.getcwd()
        dL1 = "Your_BOM_File.xlsx"
        dL2 = "Your_Feeder_File.xlsx"

        df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM_data")
        dfs1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder_data")
        df2 = pd.read_excel("FeederVerify.xlsx", sheet_name="Side")
        df3 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM and Feeder Compare")
        df4 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder and BOM Compare")
        dfs21 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder_data", usecols=['Location','F_Part_No','FeederName','Type','Size','FeedPitch','Part Height','Status','QTY','Side','ModelName','F_Ref_List','Feeder Reference'])
        dfs22 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM_data", usecols=['B_Ref_List','B_Part_No','Long Des'])
        dfs3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_BAF")
        dbf3 = pd.read_excel("FeederVerify.xlsx", sheet_name="Verify_data_FAB")

        dfsg21 = dfs21[dfs21['Feeder Reference'].duplicated()]
        dfsg22 = dfs22[dfs22['B_Ref_List'].duplicated()]
        dfsg31 = dfs3[dfs3['BOM and Feeder Compare'].str.contains('Miss_Match')]
        dfsg32 = dbf3[dbf3['Feeder and BOM Compare'].str.contains('Miss_Match')]

        self.print_to_log('\n')
        self.print_to_log(f"Date and Time: {formatted_datetime}")
        self.print_to_log('\n')

        rc = len(df1)
        rc1 = len(dfs1)

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log('BOM NAME:', dL1)
        self.print_to_log(Chd, '\\__BOM__\\', dL1)
        self.print_to_log('\n')
        self.print_to_log('FeederSetup NAME:', dL2)
        self.print_to_log(Chd, '\\__FeederSetup__\\', dL2)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("BOM Count:", rc)
        self.print_to_log('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx, Sheetname=BOM_data')
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Feeder Count:", rc1)
        self.print_to_log('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx, Sheetname=Feedercol')
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("BOT & TOP Count:")
        self.print_to_log(df2)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Compare Count:")
        self.print_to_log(df3)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Compare Count:")
        self.print_to_log(df4)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Feeder duplicate Reference")
        self.print_to_log(dfsg21)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("BOM duplicate Reference")
        self.print_to_log(dfsg22)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Miss Match Row BOM to Feeder")
        self.print_to_log(dfsg31)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~\n")

        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.print_to_log("Miss Match Row Feeder to BOM")
        self.print_to_log(dfsg32)
        self.print_to_log("~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")

    time.sleep(2)
    
if __name__ == "__main__":
    root = tk.Tk()
    app = FeederSetupApp(root)
    root.mainloop()

# Assuming feeder verification is completed
feeder_verification_completed = True

if feeder_verification_completed:
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    messagebox.showinfo("Feeder Verification", "Feeder verification has been completed!")

##########################################################################################################################################

##########################################################################################################################################

# Process to Next Upload if Match count ok

os.getcwd()

Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
Chd = os.getcwd()

for i in range(100):
    row = "="*i + ">"
    sys.stdout.write("%s\r %d%%\r" %(row, i + 1))
    sys.stdout.flush()
    time.sleep(0.1)

print("\033[1;92;4mFeederSetup___Uploading in Progress $ PROCESS $\033[0m")

print('\n')

# Load Excel sheets into pandas dataframes
dfmc1 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM and Feeder Compare")
dfmc2 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder and BOM Compare")
dfmc3 = pd.read_excel("FeederVerify.xlsx", sheet_name="BOM Count")
dfmc4 = pd.read_excel("FeederVerify.xlsx", sheet_name="Feeder Count")
#df3 = pd.read_excel('excel_sheet3.xlsx')

print(dfmc1)
print(dfmc2)
print(dfmc3)
print(dfmc4)

# Function to get numeric value safely
def get_numeric_value(df, index, column):
    try:
        return df.loc[index, column]
    except KeyError:
        return None

# Get numeric inputs from specific cells in the "count" column
num1_index0 = dfmc1.loc[0, 'count']
num1_index1 = get_numeric_value(dfmc1, 1, 'count')

num2_index0 = dfmc2.loc[0, 'count']
num2_index1 = get_numeric_value(dfmc2, 1, 'count')

num3_index0 = dfmc3.loc[0, 'BOM_Data Ref, Count']  # Note: Case-sensitive column name
num3_index1 = get_numeric_value(dfmc3, 1, 'BOM_Data Ref, Count')

num4_index0 = dfmc4.loc[0, 'Feeder_Data Ref, Count']  # Note: Case-sensitive column name
num4_index1 = get_numeric_value(dfmc4, 1, 'Feeder_Data Ref, Count')

# Compare the numeric values
if num1_index0 == num2_index0 == num3_index0 == num4_index0 and \
(num1_index1 is None or num1_index1 == num2_index1 == num3_index1 == num4_index1):
    print("Numeric values are the same. Proceeding to the next line of code.")

    # Print the values
    print("\nValues at index 0:")
    print("num1:", num1_index0)
    print("num2:", num2_index0)
    print("num3:", num3_index0)
    print("num4:", num4_index0)

    if num1_index1 is not None:
        print("\nValues at index 1:")
        print("num1:", num1_index1)
        print("num2:", num2_index1)
        print("num3:", num3_index1)
        print("num4:", num4_index1)

    else:
        print("Numeric values are not the same. Cannot proceed.")

##########################################################################################################################################

##########################################################################################################################################

#CREATE & SEPRATE FEEDER LOADING LIST DATA

    os.getcwd()

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    df1 = pd.read_excel("FeederVerify.xlsx", sheet_name="Upload_data")

    df1.sort_values(by='Side', inplace=True, ascending=True)
    df2_1 = df1
#del.TOP
    df1 = df1[df1["Side"].str.contains("TOP")==False]
    df2 = df1[df1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEXIII-2|AIMEXIII-3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df3 = df1[df1["ModelName"].str.contains("NXT|AIMEX3|AIMEXIII-1|AIMEXIII-3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df4 = df1[df1["ModelName"].str.contains("NXT|AIMEX2|AIMEXIII-1|AIMEXIII-2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df4.sort_values(by='Feeder Location', inplace=True, ascending=True)
#del.BOT
    df2_1 = df2_1[df2_1["Side"].str.contains("BOT")==False]
    df2_2 = df2_1[df2_1["ModelName"].str.contains("AIMEX2|AIMEX3|AIMEXIII-2|AIMEXIII-3|AIMEX-IIIC_2|AIMEX-IIIC_3")==False]
    df2_2.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_3 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX3|AIMEXIII-1|AIMEXIII-3|AIMEX-IIIC_1|AIMEX-IIIC_3")==False]
    df2_3.sort_values(by='Feeder Location', inplace=True, ascending=True)
    df2_4 = df2_1[df2_1["ModelName"].str.contains("NXT|AIMEX2|AIMEXIII-1|AIMEXIII-2|AIMEX-IIIC_1|AIMEX-IIIC_2")==False]
    df2_4.sort_values(by='Feeder Location', inplace=True, ascending=True)

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data.xlsx") as writer:
        df2.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df3.to_excel(writer, sheet_name="AMX2_B", index=False)
        df4.to_excel(writer, sheet_name="AMX3_B", index=False)
    
        df2_2.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df2_3.to_excel(writer, sheet_name="AMX2_T", index=False)
        df2_4.to_excel(writer, sheet_name="AMX3_T", index=False)

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#Upload data to merge and del side and Module
    
    os.getcwd()

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    df1 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_B")
    df1["Remarks"] = df1['Side'].astype(str) +"--"+ df1['ModelName']
    del df1['Side']
    del df1['ModelName']

    df2 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_B")
    df2["Remarks"] = df2['Side'].astype(str) +"--"+ df2['ModelName']
    del df2['Side']
    del df2['ModelName']

    df3 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_B")
    df3["Remarks"] = df3['Side'].astype(str) +"--"+ df3['ModelName']
    del df3['Side']
    del df3['ModelName']

    df4 = pd.read_excel("Upload-Data.xlsx", sheet_name="NXT&AMX1_T")
    df4["Remarks"] = df4['Side'].astype(str) +"--"+ df4['ModelName']
    del df4['Side']
    del df4['ModelName']

    df5 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX2_T")
    df5["Remarks"] = df5['Side'].astype(str) +"--"+ df5['ModelName']
    del df5['Side']
    del df5['ModelName']

    df6 = pd.read_excel("Upload-Data.xlsx", sheet_name="AMX3_T")
    df6["Remarks"] = df6['Side'].astype(str) +"--"+ df6['ModelName']
    del df6['Side']
    del df6['ModelName']

    with pd.ExcelWriter("D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx") as writer:
        df1.to_excel(writer, sheet_name="NXT&AMX1_B", index=False)
        df2.to_excel(writer, sheet_name="AMX2_B", index=False)
        df3.to_excel(writer, sheet_name="AMX3_B", index=False)
        df4.to_excel(writer, sheet_name="NXT&AMX1_T", index=False)
        df5.to_excel(writer, sheet_name="AMX2_T", index=False)
        df6.to_excel(writer, sheet_name="AMX3_T", index=False)

##########################################################################################################################################

##########################################################################################################################################

    #bil10 = pyfiglet.figlet_format("Feeder Loading List Progress", width = 200)
    print('\n')
    print("\033[1;92;4m******Feeder Loading List Progress******\033[0m")
    print('\n')
##########################################################################################################################################

##########################################################################################################################################

#Feeder List change

    shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Database_File/SMT_FeederSetup/Line X Sample.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx')

        ##BOT FEEDER LIST

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'NXT&AMX1_B'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'NXT'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "NXT&AMX1_B", "NXT", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX2_B'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'AIMEX 2'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX2_B", "AIMEX 2", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX3_B'

        # Destination Excel file 
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx'
    destination_sheet_name = 'AIMEX 3'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample B.xlsx", "AMX3_B", "AIMEX 3", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

        ##TOP FEEDER LIST

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'NXT&AMX1_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'NXT'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "NXT&AMX1_T", "NXT", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX2_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'AIMEX 2'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX2_T", "AIMEX 2", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

        # Source Excel file
    source_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data-TB.xlsx'
    source_sheet_name = 'AMX3_T'

        # Destination Excel file
    destination_file = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx'
    destination_sheet_name = 'AIMEX 3'

    def copy_data_with_offset(source_file, destination_file, source_sheet_name, destination_sheet_name, offset):
            # Load the source workbook
            source_workbook = openpyxl.load_workbook(source_file)
            source_sheet = source_workbook[source_sheet_name]

            # Load the destination workbook
            destination_workbook = openpyxl.load_workbook(destination_file)
            destination_sheet = destination_workbook[destination_sheet_name]

            # Iterate through the source sheet and copy data to the destination sheet with an offset
            for row_index, row in enumerate(source_sheet.iter_rows(values_only=True), start=1):
                # Offset the row index by the specified offset
                destination_row = row_index + offset

                # Copy data to the destination sheet
                for col_index, value in enumerate(row, start=1):
                    destination_sheet.cell(row=destination_row, column=col_index, value=value)

            # Save the changes to the destination workbook
            destination_workbook.save(destination_file)

        # Example usage:
    copy_data_with_offset("Upload-Data-TB.xlsx", "Line X Sample T.xlsx", "AMX3_T", "AIMEX 3", offset=5)
        # Example usage:copy_data("source_workbook.xlsx", "destination_workbook.xlsx", "Sheet1", "Sheet2")

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

##########################################################################################################################################

#CREATEBACKUPFOLDER

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    yourfolder1 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\BOM"

    if not os.path.isdir(yourfolder1):
        print('Folder Not Exist')
        os.makedirs(yourfolder1)

    yourfolder2 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\FeederSetup"

    if not os.path.isdir(yourfolder2):
        print('Folder Not Exist')
        os.makedirs(yourfolder2)

    yourfolder3 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\Upload"

    if not os.path.isdir(yourfolder3):
        print('Folder Not Exist')
        os.makedirs(yourfolder3)

    yourfolder4 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\Verified"

    if not os.path.isdir(yourfolder4):
        print('Folder Not Exist')
        os.makedirs(yourfolder4)
        
    yourfolder5 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\AVL & Polarity Check"

    if not os.path.isdir(yourfolder5):
        print('Folder Not Exist')
        os.makedirs(yourfolder5)
        
    yourfolder6 = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#_Output\PartNumber"

    if not os.path.isdir(yourfolder6):
        print('Folder Not Exist')
        os.makedirs(yourfolder6)

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM')
    Chd = os.getcwd()

    file_path = 'BOM_List_OP.xlsx'
    directory_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/BOM'

    print(os.path.isfile(file_path))
    print(os.path.isfile(directory_path))

    try:
        if os.path.isfile(file_path):
            dt_H1 = pd.read_csv(file_path)

    except ValueError:
        if os.path.exists("BOM_List_OP.xlsx"):
            os.remove("BOM_List_OP.xlsx")
    else:
        print("The file does not exist")

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')
    Chd = os.getcwd()

    if os.path.exists("Feeder_List_OPT.xlsx"):
        os.remove("Feeder_List_OPT.xlsx")
    else:
        print("The file does not exist")

    if os.path.exists("Feeder_List_OPB.xlsx"):
        os.remove("Feeder_List_OPB.xlsx")
    else:
        print("The file does not exist")

    ##########################################################################################################################################

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')

    yourfolder = r"D:\NX_BACKWORK\Feeder Setup_PROCESS\#Output\Line_X"

    if not os.path.isdir(yourfolder):
        print('Folder Not Exist')
        os.makedirs(yourfolder)

    # Get the current working directory
    os.getcwd()

    # Change directory to the location of your Excel files
    os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload')

    # Get the current working directory
    Chd = os.getcwd()

    # Load the workbooks
    workbook_T = openpyxl.load_workbook('Line X Sample T.xlsx')
    workbook_B = openpyxl.load_workbook('Line X Sample B.xlsx')

    # Specify the worksheet names
    worksheet_names = ['NXT', 'AIMEX 2', 'AIMEX 3']

    # Input values for cell B3 and Revision A1
    #print('\n')
    #value_B3 = input("\033[93mEnter Feeder Name (12 characters): \033[0m").strip()[:12]  # Take only the first 12 characters and remove extra spaces
    #print('\n')
    #Revision = input("\033[93mEnter Revision A1: \033[0m")

    # Iterate over each workbook
    for workbook, workbook_name in [(workbook_T, 'Line X Sample T.xlsx'), (workbook_B, 'Line X Sample B.xlsx')]:
        if 'T' in workbook_name:
            location = 'T'
        elif 'B' in workbook_name:
            location = 'B'
        else:
            location = 'T/B'

        # Iterate over each sheet in the workbook
        for sheet_name in worksheet_names:
            # Select the worksheet
            worksheet = workbook[sheet_name]

            # Iterate over all rows in the worksheet
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                # Check if the row is empty
                if all(cell.value is None for cell in row):
                    # If the row is empty, hide it
                    worksheet.row_dimensions[row[0].row].hidden = True

            # Combine the input values for B3 & K4 cell
            worksheet['B3'] = value_B3 + " " + location + " " + Revision
            worksheet['K4'] = dLine123

    # Save the workbooks
    workbook_T.save('Line X Sample T.xlsx')
    workbook_B.save('Line X Sample B.xlsx')

    #shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Upload-Data.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Upload-Data.xlsx')
    #shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/PartMaster.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/PartMaster.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/MODEL.mdb', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/MODEL.mdb')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample T.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample T.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Upload/Line X Sample B.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/Line X Sample B.xlsx')

    ##########################################################################################################################################

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified')
    Chd = os.getcwd()

    #shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/BOM_List_OP.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/BOM_List_OP.xlsx')
    #shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederSetup.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/FeederSetup.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/FeederVerify.xlsx', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/FeederVerify.xlsx')
    shutil.copyfile('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Verified/AVL.csv', 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X/AVL.CSV')

    ##########################################################################################################################################

    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output/Line_X')
    Chd = os.getcwd()

    #src_1 = 'Upload-Data.xlsx'
    #os.rename(src_1, dL2 +"_UD"+".xlsx")

    #src_2 = 'PartMaster.xlsx'
    #os.rename(src_2, dL2 +"_PM"+".xlsx")

    src_3 = 'MODEL.mdb'
    os.rename(src_3, dL2 +"_PM-Model"+".mdb")

    #src_4 = 'BOM_List_OP.xlsx'
    #os.rename(src_4, dL1 +"_BOM"+".xlsx")

    #src_5 = 'FeederSetup.xlsx'
    #os.rename(src_5, dL2 +"_FS"+".xlsx")

    src_6 = 'FeederVerify.xlsx'
    os.rename(src_6, dL2 +"_FV"+".xlsx")

    src_7 = 'AVL.csv'
    os.rename(src_7, dL1 +"_AVL"+".csv")

    src_8 = 'Line X Sample T.xlsx'
    os.rename(src_8, dL1 +"_T_"+ Revision +".xlsx")

    src_9 = 'Line X Sample B.xlsx'
    os.rename(src_9, dL1 +"_B_"+ Revision +".xlsx")

    # Specify the current name of the folder
    cfn1 = "Line_X"

    # Rename the folder
    os.getcwd()
    Chd= os.chdir('D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output')
    Chd = os.getcwd()
    os.rename(cfn1, dL1 +"-"+ Revision)

    print(f"\033[92mFolder '{cfn1}' renamed successfully to '{dL1}'.\033[0m")
    
    # Function to move a folder to a specified destination
    def move_folder(src, dst):
        try:
            shutil.move(src, dst)
            messagebox.showinfo("Success", f"Folder '{os.path.basename(src)}' moved successfully to '{dst}'")
        except Exception as e:
            messagebox.showerror("Error", f"Error occurred while moving folder:\n{e}")

    # Function to create a new customer name
    def create_customer_name(line):
        new_customer = simpledialog.askstring("Create New Customer", f"Enter the name of the new customer for '{line}':")
        if new_customer:
            line_path = os.path.join(destination_path, line)
            new_customer_path = os.path.join(line_path, new_customer)
            os.makedirs(new_customer_path, exist_ok=True)
            return new_customer
        else:
            return None
        
    import tkinter as tk
    from tkinter import ttk, messagebox, simpledialog
    import os
    import shutil

    # Define the root directory
    root_dir = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#Output'

    # Define the destination path
    destination_path = 'D:/NX_BACKWORK/Feeder Setup_PROCESS/#_Feeder Loading Line List'

    # Get a list of lines from the destination path
    lines = [name for name in os.listdir(destination_path) if os.path.isdir(os.path.join(destination_path, name))]

    # Define the layout for the PySimpleGUI window
    class MoveFolderApp:
        def __init__(self, master):
            self.master = master
            self.master.title("Move Folder")

            # Folder selection
            tk.Label(master, text="Select the folder to move:").grid(row=0, column=0, sticky="w")
            self.folder_listbox = tk.Listbox(master, height=6, width=60)
            self.folder_listbox.grid(row=1, column=0, padx=10, pady=5)
            for folder in os.listdir(root_dir):
                self.folder_listbox.insert(tk.END, folder)
            self.folder_listbox.bind("<<ListboxSelect>>", self.on_folder_select)

            # Line selection
            tk.Label(master, text="Select the line:").grid(row=2, column=0, sticky="w")
            self.line_listbox = tk.Listbox(master, height=3, width=60)
            self.line_listbox.grid(row=3, column=0, padx=10, pady=5)
            for line in lines:
                self.line_listbox.insert(tk.END, line)
            self.line_listbox.bind("<<ListboxSelect>>", self.on_line_select)

            # Customer selection and creation
            tk.Label(master, text="Select or create a customer name:").grid(row=4, column=0, sticky="w")
            self.customer_listbox = tk.Listbox(master, height=6, width=60)
            self.customer_listbox.grid(row=5, column=0, padx=10, pady=5)

            # Frame to hold action buttons in a row
            button_frame = tk.Frame(master)
            button_frame.grid(row=6, column=0, columnspan=2, pady=10)

            self.new_customer_button = tk.Button(button_frame, text="New Customer", command=self.create_customer)
            self.new_customer_button.pack(side="left", padx=5)

            self.move_button = tk.Button(button_frame, text="Move", command=self.move_folder)
            self.move_button.pack(side="left", padx=5)

            self.cancel_button = tk.Button(button_frame, text="Cancel", command=self.master.quit)
            self.cancel_button.pack(side="left", padx=5)


        def on_folder_select(self, event):
            self.selected_folder = self.folder_listbox.get(tk.ACTIVE)

        def on_line_select(self, event):
            self.selected_line = self.line_listbox.get(tk.ACTIVE)
            path = os.path.join(destination_path, self.selected_line)
            customers = os.listdir(path) if os.path.exists(path) else []
            self.customer_listbox.delete(0, tk.END)
            for customer in customers:
                self.customer_listbox.insert(tk.END, customer)

        def create_customer(self):
            try:
                line = self.line_listbox.get(tk.ACTIVE)
                if not line:
                    messagebox.showerror("Error", "Please select a line first.")
                    return
                new_customer = simpledialog.askstring("New Customer", f"Enter the name of the new customer for '{line}':")
                if new_customer:
                    new_customer_path = os.path.join(destination_path, line, new_customer)
                    os.makedirs(new_customer_path, exist_ok=True)
                    self.customer_listbox.insert(tk.END, new_customer)
            except Exception as e:
                messagebox.showerror("Error", str(e))

        def move_folder(self):
            try:
                folder = self.folder_listbox.get(tk.ACTIVE)
                line = self.line_listbox.get(tk.ACTIVE)
                customer = self.customer_listbox.get(tk.ACTIVE)
                if not folder:
                    messagebox.showerror("Error", "Please select a folder to move.")
                    return
                if not line:
                    messagebox.showerror("Error", "Please select a line.")
                    return
                if not customer:
                    messagebox.showerror("Error", "Please select or create a customer.")
                    return

                src_path = os.path.join(root_dir, folder)
                dst_path = os.path.join(destination_path, line, customer, folder)
                shutil.move(src_path, dst_path)
                messagebox.showinfo("Success", f"Folder '{folder}' moved to '{dst_path}'")
            except Exception as e:
                messagebox.showerror("Error", f"Error occurred while moving folder: {e}")
                
# Launch the app
if __name__ == "__main__":
    root = tk.Tk()
    app = MoveFolderApp(root)
    root.mainloop()

    time.sleep (5)

    print('\n')
    print('\033[92;3mFeeder Setup Generation Complete\033[0m')
    print('\n')
    print('\033[92;3mBOM and Feeder Verfication Found OK\033[0m')
    print('\n')

    # Assuming feeder verification is completed
    Feeder_List_Generation_Completed = True

    if Feeder_List_Generation_Completed:
        root = tk.Tk()
        root.withdraw()  # Hide the main window

        messagebox.showinfo("Feeder Loading List", "Feeder Loading List has been Generated!")

else:
    # Abort the process
    print("Counts are different. Aborting the process.")
    # Show error message
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    error_message = "Counts are different. Aborting the Feeder Loading Process\nCheck the Feeder Verfied for Miss_Match."
    messagebox.showerror("Error", error_message)
    # Exit the script
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
sys.exit() #FeederSetup X

#pyinstaller -F --onefile --console --name Feeder Verifier --icon=SYRMA.ico FeederVerfier.py
